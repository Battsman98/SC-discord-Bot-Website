import asyncio
import difflib
import hashlib
import html
import json
import logging
import os
import queue
import re
import secrets
import threading
import aiohttp
from io import BytesIO
import time
from contextlib import asynccontextmanager, suppress
from dataclasses import asdict, is_dataclass
from pathlib import Path
from typing import Any

from fastapi import Depends, FastAPI, File, Form, Header, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from src.bot import (
    CZ_TIMER_DEFINITIONS,
    CZ_TIMERS_CACHE_KEY,
    EXEC_OVERRIDE_CACHE_KEY,
    _has_mining_multi_separator,
    _mining_multi_search_terms,
    _mining_space_search_terms,
    _mining_term_signatures,
    _shared_mining_signatures,
    _unique_preserve_order,
    add_community_mining_location,
    apply_community_mining_locations,
    get_cz_dashboard_timers,
)
from src.cache import AUDIT_ACTION_TYPES, SQLiteCache
from src.config import Settings
from src.security import SlidingWindowLimiter, install_secret_redaction
from src.sources.base import ItemLocatorResult
from src.sources.citizen_updates import CitizenUpdatesSource
from src.sources.p4k_inventory import P4KInventoryCatalog
from src.sources.registry import SourceRegistry, build_default_registry
from src.sources.warbonds import WarbondTrackerSource
from src.timers import (
    calculate_countdown_end_unix,
    calculate_cycle_start_from_phase,
    calculate_exec_hangar_status,
    fetch_exec_cycle_start_unix,
)
from src.web_auth import (
    OAUTH_STATE_COOKIE_NAME,
    SESSION_COOKIE_NAME,
    build_discord_authorize_url,
    current_user_from_request,
    discord_auth_configured,
    encode_session,
    exchange_discord_code,
    fetch_web_user,
    human_verification_configured,
    oauth_state,
    session_secret,
    verify_human,
)


ROOT_DIR = Path(__file__).resolve().parents[1]
WEB_DIR = ROOT_DIR / "web"
P4K_INVENTORY_CATALOG = P4KInventoryCatalog(ROOT_DIR / "data" / "p4k_inventory_snapshot.json")
COMMANDS_PATH = ROOT_DIR / "docs" / "commands.md"
SHIP_LOANERS = {
    "600i explorer": ["Cyclone"],
    "600i executive": ["Cyclone"],
    "890 jump": ["85x"],
    "arrastra": ["Anvil Arrow", "Argo MOLE", "MISC Prospector"],
    "rsi arrastra": ["Anvil Arrow", "Argo MOLE", "MISC Prospector"],
    "carrack": ["C8 Pisces", "URSA Rover"],
    "carrack expedition": ["C8 Pisces", "URSA Rover"],
    "carrack w c8x": ["C8X Pisces Expedition", "URSA Rover"],
    "carrack expedition w c8x": ["C8X Pisces Expedition", "URSA Rover"],
    "centurion": ["Aurora MR"],
    "constellation andromeda": ["P-52 Merlin"],
    "constellation aquila": ["P-52 Merlin", "URSA Rover"],
    "constellation phoenix": ["P-72 Archimedes", "Lynx Rover"],
    "constellation phoenix emerald": ["P-72 Archimedes", "Lynx Rover"],
    "crucible": ["Constellation Andromeda"],
    "csv sm": ["Aurora MR"],
    "cyclone": ["Aurora MR"],
    "cyclone aa": ["Aurora MR"],
    "cyclone mt": ["Aurora MR"],
    "cyclone rc": ["Aurora MR"],
    "cyclone rn": ["Aurora MR"],
    "cyclone tr": ["Aurora MR"],
    "dragonfly": ["Aurora MR"],
    "e1 spirit": ["A1 Spirit"],
    "endeavor": ["Starfarer", "Cutlass Red"],
    "expanse": ["Prospector", "Reliant Kore"],
    "fury": ["Aurora MR"],
    "fury lx": ["Aurora MR"],
    "fury mx": ["Aurora MR"],
    "g12": ["Lynx"],
    "g12 a": ["Lynx"],
    "g12 r": ["Lynx"],
    "galaxy": ["Anvil Carrack"],
    "rsi galaxy": ["Anvil Carrack"],
    "genesis starliner": ["Hercules C2"],
    "hull d": ["Hull C", "Hercules C2"],
    "hull e": ["Hull C", "Hercules C2"],
    "idris m": ["F7C-M Super Hornet", "MPUV Passenger"],
    "idris p": ["F7C-M Super Hornet", "MPUV Passenger"],
    "javelin": ["Idris-P", "MPUV Cargo"],
    "kraken": ["Polaris", "Ironclad Assault", "Buccaneer"],
    "kraken privateer": ["Polaris", "Ironclad", "Buccaneer"],
    "liberator": ["Ironclad Assault", "F7C-M Super Hornet"],
    "legionnaire": ["Vanguard Hoplite"],
    "lynx": ["Aurora MR"],
    "mantis": ["Aurora LN"],
    "merchantman": ["Hull C", "Defender", "Hercules C2"],
    "banu merchantman": ["Hull C", "Defender", "Hercules C2"],
    "mole": ["Prospector"],
    "mpuv tractor": ["Aurora MR"],
    "mxc": ["Aurora MR"],
    "mule": ["Aurora MR"],
    "nautilus": ["Polaris", "Avenger Titan"],
    "nova": ["Aurora MR"],
    "nox": ["Aurora MR"],
    "odin": ["Idris-P"],
    "odyssey": ["Carrack", "Reliant Kore"],
    "orion": ["Prospector", "Mole"],
    "pioneer": ["Caterpillar", "Nomad"],
    "pitbull": ["Aurora MR"],
    "pulse": ["Aurora MR"],
    "pulse lx": ["Aurora MR"],
    "railen": ["Hercules C2", "Syulen"],
    "ranger cv": ["Cyclone"],
    "ranger rc": ["Cyclone RC"],
    "ranger tr": ["Cyclone TR"],
    "redeemer": ["Arrow"],
    "srv": ["Aurora LN"],
    "storm": ["Aurora MR"],
    "storm aa": ["Aurora MR"],
    "storm variants": ["Aurora MR"],
    "stv": ["Aurora MR"],
    "utv": ["Aurora MR"],
    "vulcan": ["Starfarer"],
    "x1": ["Aurora MR"],
    "x1 force": ["Aurora MR"],
    "x1 velocity": ["Aurora MR"],
    "zeus mk ii mr": ["Zeus Mk II ES"],
}
SHIP_DISPLAY_PREFIXES = (
    "Aegis ",
    "Anvil ",
    "Aopoa ",
    "Argo ",
    "Banu ",
    "Consolidated Outland ",
    "Crusader ",
    "Drake ",
    "Esperia ",
    "Gatac ",
    "Greycat ",
    "Kruger ",
    "MISC ",
    "Mirai ",
    "Origin ",
    "RSI ",
    "Tumbril ",
)


class AppState:
    settings: Settings
    cache: SQLiteCache
    sources: SourceRegistry
    updates: CitizenUpdatesSource
    warbonds: WarbondTrackerSource
    item_catalog_task: asyncio.Task | None
    warbond_task: asyncio.Task | None
    scanner_gate: "InventoryScannerGate"
    request_limiter: SlidingWindowLimiter


class InventoryScannerGate:
    """Bound live OCR concurrency while keeping each admitted scan user-bound."""

    def __init__(self, worker_count: int = 2, capacity: int = 4) -> None:
        self.worker_count = max(1, worker_count)
        self.capacity = max(self.worker_count, capacity)
        self._workers = asyncio.Semaphore(self.worker_count)
        self._state_lock = asyncio.Lock()
        self._admitted_users: set[int] = set()

    @asynccontextmanager
    async def admit(self, user_id: int):
        scan_id = secrets.token_urlsafe(12)
        async with self._state_lock:
            if user_id in self._admitted_users:
                raise HTTPException(
                    status_code=409,
                    detail="This account already has an inventory scan in progress.",
                )
            if len(self._admitted_users) >= self.capacity:
                raise HTTPException(
                    status_code=429,
                    detail="The inventory scanner is busy. Try again in a few seconds.",
                    headers={"Retry-After": "2"},
                )
            self._admitted_users.add(user_id)

        queued_at = time.perf_counter()
        try:
            async with self._workers:
                yield scan_id, round((time.perf_counter() - queued_at) * 1000)
        finally:
            async with self._state_lock:
                self._admitted_users.discard(user_id)


class MiningCommunityRequest(BaseModel):
    material: str = Field(min_length=1)
    system: str = Field(min_length=1)
    location_type: str = Field(min_length=1)
    location: str = Field(min_length=1)
    reported_by: str = "Website"


class RefineryOrderRequest(BaseModel):
    label: str = Field(min_length=1, max_length=120)
    material: str | None = Field(default=None, max_length=120)
    quantity: float | None = Field(default=None, ge=0)
    refinery: str | None = Field(default=None, max_length=120)
    method: str | None = Field(default=None, max_length=120)
    location: str | None = Field(default=None, max_length=120)
    crew: str | None = Field(default=None, max_length=500)
    notes: str | None = Field(default=None, max_length=2000)
    completes_at: int = Field(gt=0)


class RefineryOrderStatusRequest(BaseModel):
    status: str


class ExecOverrideRequest(BaseModel):
    phase: str
    remaining_minutes: int = Field(gt=0)
    corrected_by: str = "Website"


class CZTimerRequest(BaseModel):
    timer: str
    started_minutes_ago: int = Field(default=0, ge=0)


class BlueprintOwnershipRequest(BaseModel):
    name: str = Field(min_length=1)
    category: str | None = None
    source_name: str | None = None
    source_url: str | None = None


class BlueprintTextImportRequest(BaseModel):
    text: str = Field(min_length=1)


class ShipOwnershipRequest(BaseModel):
    name: str = Field(min_length=1)
    ownership_type: str
    manufacturer: str | None = None
    role: str | None = None
    vehicle_type: str | None = None
    size: str | None = None
    status: str | None = None
    cargo_capacity: float | None = None
    source_name: str | None = None
    source_url: str | None = None
    image_url: str | None = None
    notes: str | None = None
    quantity: int | None = Field(default=None, ge=1, le=999)
    increment: bool = False


class RsiPledgeImportRequest(BaseModel):
    pages: list[str] = Field(default_factory=list)
    candidates: list[str] = Field(default_factory=list)


class InventoryItemRequest(BaseModel):
    name: str = Field(min_length=1)
    category: str | None = None
    location: str = Field(min_length=1)
    quantity: int = Field(default=1, ge=0)
    quality: float | None = Field(default=None, ge=0)
    item_type: str | None = None
    item_size: str | None = None
    volume_scu: float | None = Field(default=None, ge=0)
    notes: str | None = None


class InventoryTransferRequest(BaseModel):
    location: str = Field(min_length=1)


class InventoryClearRequest(BaseModel):
    location: str | None = None


class InventoryTextImportRequest(BaseModel):
    text: str = Field(min_length=1)
    default_location: str | None = None
    default_category: str | None = None
    default_item_type: str | None = None
    scanner_mode: bool = False
    min_score: float = Field(default=0.72, ge=0, le=1)
    exclude_words: str | None = None


@asynccontextmanager
async def lifespan(app: FastAPI):
    settings = Settings.from_env(require_discord_token=False)
    install_secret_redaction()
    cache = await SQLiteCache.create(settings.database_path)
    sources = await build_default_registry(settings, cache)
    updates = CitizenUpdatesSource(settings, cache)
    warbonds = WarbondTrackerSource(cache, settings.http_timeout_seconds)
    app.state.game_assist = AppState()
    app.state.game_assist.settings = settings
    app.state.game_assist.cache = cache
    app.state.game_assist.sources = sources
    app.state.game_assist.updates = updates
    app.state.game_assist.warbonds = warbonds
    app.state.game_assist.scanner_gate = InventoryScannerGate(
        worker_count=int(os.getenv("INVENTORY_SCANNER_WORKERS", "2")),
        capacity=int(os.getenv("INVENTORY_SCANNER_CAPACITY", "4")),
    )
    app.state.game_assist.request_limiter = SlidingWindowLimiter(
        settings.web_rate_limit_per_minute, 60
    )
    app.state.game_assist.item_catalog_task = asyncio.create_task(
        _item_catalog_maintenance_loop(sources),
        name="item-catalog-maintenance",
    )
    app.state.game_assist.warbond_task = asyncio.create_task(
        _warbond_maintenance_loop(warbonds),
        name="warbond-maintenance",
    )
    try:
        await asyncio.to_thread(_initialize_rapid_ocr_pool)
        await asyncio.gather(*(
            asyncio.to_thread(_warm_rapid_title_ocr)
            for _ in range(_RAPID_TITLE_OCR_POOL_SIZE)
        ))
        from PIL import Image
        warm_frame = Image.new("RGB", (1280, 720), "black")
        warm_output = BytesIO()
        warm_frame.save(warm_output, format="PNG")
        warm_image_data = warm_output.getvalue()
        await asyncio.gather(*(
            asyncio.to_thread(
                _read_calibrated_inventory_title,
                warm_image_data,
                _DEFAULT_INVENTORY_TITLE_BOX,
            )
            for _ in range(_RAPID_TITLE_OCR_POOL_SIZE)
        ))
        # Build the local item-catalog indexes during startup instead of making
        # the first scanner request pay the cold-load cost.
        await sources.lookup_inventory_items("inventory scanner warmup", limit=1)
    except Exception:
        pass
    try:
        yield
    finally:
        app.state.game_assist.item_catalog_task.cancel()
        app.state.game_assist.warbond_task.cancel()
        with suppress(asyncio.CancelledError):
            await app.state.game_assist.item_catalog_task
        with suppress(asyncio.CancelledError):
            await app.state.game_assist.warbond_task
        await warbonds.close()
        await updates.close()
        await sources.close()
        await cache.close()


app = FastAPI(
    title="Game Assist Web",
    description="Website companion API for the Star Citizen Discord bot.",
    version="0.1.0",
    lifespan=lifespan,
)
_RAPID_OCR = None
_RAPID_OCR_LOCK = threading.Lock()
_RAPID_OCR_POOL_SIZE = 1
_RAPID_OCR_POOL: queue.LifoQueue[Any] = queue.LifoQueue(maxsize=_RAPID_OCR_POOL_SIZE)
_RAPID_OCR_POOL_READY = False
_RAPID_TITLE_OCR_POOL_SIZE = max(1, int(os.getenv("INVENTORY_SCANNER_WORKERS", "2")))
_RAPID_TITLE_OCR_POOL: queue.LifoQueue[Any] = queue.LifoQueue(maxsize=_RAPID_TITLE_OCR_POOL_SIZE)
_RAPID_TITLE_OCR_POOL_READY = False
_DEFAULT_INVENTORY_TITLE_BOX = "0.300000,0.245000,0.380000,0.027500"
_INVENTORY_TITLE_FALLBACK_BOXES = (
    "0.300000,0.250000,0.380000,0.027500",
    "0.300000,0.205000,0.380000,0.027500",
    "0.300000,0.295000,0.380000,0.027500",
    "0.300000,0.340000,0.380000,0.027500",
    "0.300000,0.160000,0.380000,0.027500",
    "0.300000,0.385000,0.380000,0.027500",
    "0.300000,0.430000,0.380000,0.027500",
    "0.300000,0.475000,0.380000,0.027500",
    "0.300000,0.520000,0.380000,0.027500",
    "0.300000,0.565000,0.380000,0.027500",
    "0.300000,0.610000,0.380000,0.027500",
)
VISITOR_COOKIE_NAME = "sc_companion_visitor"
_STATE_CHANGING_METHODS = {"POST", "PUT", "PATCH", "DELETE"}


@app.middleware("http")
async def security_controls(request: Request, call_next):
    """Apply browser, request-size, origin, and abuse protections in one place."""
    if hasattr(app.state, "game_assist"):
        settings = state().settings
        # Uvicorn resolves trusted proxy headers into request.client. Reading a raw
        # X-Forwarded-For value here would let clients rotate a spoofed rate-limit key.
        client_ip = request.client.host if request.client else "unknown"
        allowed, retry_after = state().request_limiter.allow(client_ip)
        if not allowed:
            return Response(status_code=429, headers={"Retry-After": str(retry_after)})

        content_length = request.headers.get("content-length")
        if content_length:
            try:
                if int(content_length) > settings.max_upload_bytes:
                    return Response(status_code=413)
            except ValueError:
                return Response(status_code=400)

        if request.method in _STATE_CHANGING_METHODS:
            origin = request.headers.get("origin", "").rstrip("/")
            fetch_site = request.headers.get("sec-fetch-site", "").lower()
            request_origin = f"{request.url.scheme}://{request.url.netloc}".rstrip("/")
            trusted = {request_origin, *settings.trusted_origins}
            if fetch_site == "cross-site" or (origin and origin not in trusted):
                return Response(status_code=403)

    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    response.headers["Cross-Origin-Opener-Policy"] = "same-origin"
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; base-uri 'self'; frame-ancestors 'none'; object-src 'none'; "
        "img-src 'self' data: https://cdn.discordapp.com https:; "
        "style-src 'self' 'unsafe-inline'; "
        "script-src 'self' https://challenges.cloudflare.com; "
        "frame-src https://challenges.cloudflare.com; connect-src 'self' https:"
    )
    if request.url.path.startswith(("/api/me", "/api/audit", "/auth/")):
        response.headers["Cache-Control"] = "no-store"
    return response


def state() -> AppState:
    return app.state.game_assist


async def _item_catalog_maintenance_loop(sources: SourceRegistry) -> None:
    while True:
        try:
            status = await sources.validate_item_catalog()
        except Exception:
            status = {"status": "unavailable"}
        retry_seconds = 5 * 60 if status.get("status") in {"empty", "unavailable"} else 6 * 60 * 60
        await asyncio.sleep(retry_seconds)


async def _warbond_maintenance_loop(warbonds: WarbondTrackerSource) -> None:
    while True:
        try:
            await warbonds.active(force_refresh=True)
        except Exception:
            pass
        await asyncio.sleep(15 * 60)


@app.middleware("http")
async def track_website_visitors(request: Request, call_next):
    response = await call_next(request)
    if request.url.path in {"/assets/app.js", "/assets/styles.css"}:
        response.headers["Cache-Control"] = "no-cache, max-age=0, must-revalidate"
    is_page_visit = request.method == "GET" and request.url.path == "/"
    is_activity_ping = request.method == "POST" and request.url.path == "/api/activity"
    if (is_page_visit or is_activity_ping) and response.status_code < 400 and hasattr(app.state, "game_assist"):
        visitor_id = request.cookies.get(VISITOR_COOKIE_NAME, "")
        new_visitor = re.fullmatch(r"[a-f0-9]{32}", visitor_id) is None
        if new_visitor:
            visitor_id = secrets.token_hex(16)
        visitor_hash = hashlib.sha256(visitor_id.encode("ascii")).hexdigest()
        user = current_user_from_request(request, state().settings)
        try:
            if is_page_visit:
                await state().cache.record_website_visit(visitor_hash, user.id if user else None)
            await state().cache.touch_website_activity(visitor_hash, user.id if user else None)
        except Exception:
            pass
        if new_visitor:
            response.set_cookie(
                VISITOR_COOKIE_NAME,
                visitor_id,
                max_age=365 * 24 * 60 * 60,
                httponly=True,
                secure=request.url.scheme == "https",
                samesite="lax",
            )
    return response


@app.middleware("http")
async def audit_website_action(request: Request, call_next):
    metadata = _website_audit_metadata(request.method, request.url.path)
    response = None
    error_name = None
    try:
        response = await call_next(request)
        return response
    except Exception as error:
        error_name = type(error).__name__
        raise
    finally:
        explicit_success = (
            _website_has_explicit_audit(request.method, request.url.path)
            and response is not None
            and response.status_code < 400
        )
        if metadata and not explicit_success and hasattr(app.state, "game_assist"):
            action_type, title = metadata
            user = current_user_from_request(request, state().settings)
            fields = {
                "Source": "Website",
                "User": _website_audit_user(user),
                "Method": request.method,
                "Path": request.url.path,
                "Status": response.status_code if response is not None else 500,
                "Outcome": "Success" if response is not None and response.status_code < 400 else "Failed",
            }
            safe_query = _safe_audit_query(request)
            if safe_query:
                fields["Query"] = safe_query
            if error_name:
                fields["Error"] = error_name
            try:
                await state().cache.add_audit_event(title, fields, action_type)
            except Exception:
                pass


def _website_audit_metadata(method: str, path: str) -> tuple[str, str] | None:
    if method in {"HEAD", "OPTIONS"} or path in {"/api/health", "/api/me"} or "/autocomplete/" in path:
        return None
    if path.endswith("/facets"):
        return None
    mappings = (
        ("/api/me/inventory", "inventory", "Website Inventory Action"),
        ("/api/me/blueprints", "blueprints", "Website Blueprint Collection Action"),
        ("/api/me/ships", "ships", "Website Hangar Action"),
        ("/api/ships", "ships", "Website Ship Search"),
        ("/api/mining", "mining", "Website Mining Action"),
        ("/api/commodities", "trade", "Website Commodity Search"),
        ("/api/trade", "trade", "Website Trade Route Search"),
        ("/api/blueprints", "blueprints", "Website Blueprint Search"),
        ("/api/missions", "blueprints", "Website Mission Search"),
        ("/api/items", "items", "Website Item Search"),
        ("/api/exec", "timers", "Website Executive Timer Action"),
        ("/api/cz", "timers", "Website CZ Timer Action"),
        ("/api/commands", "commands", "Website Commands Viewed"),
        ("/api/updates", "updates", "Website Updates Viewed"),
        ("/api/audit", "audit", "Website Audit Viewed"),
        ("/api/lookup", "other", "Website General Lookup"),
        ("/auth/", "authentication", "Website Authentication Action"),
    )
    return next(
        ((action_type, title) for prefix, action_type, title in mappings if path.startswith(prefix)),
        None,
    )


def _website_has_explicit_audit(method: str, path: str) -> bool:
    return method != "GET" and (
        path == "/api/mining/community"
        or path == "/api/exec/override"
        or path.startswith("/api/cz/timers")
    )
def _website_audit_user(user: Any) -> str:
    if user is None:
        return "Anonymous"
    return f"{user.display_name or user.username} ({user.id})"


def _safe_audit_query(request: Request) -> str:
    secret_names = {"code", "state", "token", "secret", "password"}
    values = [
        f"{key}={value}"
        for key, value in request.query_params.multi_items()
        if key.lower() not in secret_names
    ]
    return "&".join(values)[:500]


def require_change_admin(
    request: Request,
    x_admin_token: str | None = Header(default=None),
) -> None:
    user = current_user_from_request(request, state().settings)
    if user and user.can_manage_changes:
        return
    require_legacy_admin_token(x_admin_token)


def require_bot_admin(
    request: Request,
    x_admin_token: str | None = Header(default=None),
) -> None:
    user = current_user_from_request(request, state().settings)
    if user and user.can_manage_admin:
        return
    require_legacy_admin_token(x_admin_token)


def require_legacy_admin_token(x_admin_token: str | None = Header(default=None)) -> None:
    token = state().settings.web_admin_token
    if not token:
        raise HTTPException(status_code=401, detail="Discord login with the required permissions is needed.")
    if x_admin_token != token:
        raise HTTPException(status_code=401, detail="Discord login with the required permissions is needed.")


def require_user(request: Request):
    user = current_user_from_request(request, state().settings)
    if user is None:
        raise HTTPException(status_code=401, detail="Discord login is required.")
    return user


def encode(value: Any) -> Any:
    if is_dataclass(value):
        return encode(asdict(value))
    if isinstance(value, list):
        return [encode(item) for item in value]
    if isinstance(value, dict):
        return {str(key): encode(item) for key, item in value.items()}
    return value


def not_found(message: str) -> None:
    raise HTTPException(status_code=404, detail=message)


@app.get("/api/health")
async def health() -> dict[str, Any]:
    settings = state().settings
    catalog = await state().sources.item_catalog_status()
    return {
        "status": "online",
        "revision": os.getenv("RENDER_GIT_COMMIT", "local")[:12],
        "discord_auth_enabled": discord_auth_configured(settings),
        "inventory_scanner": {
            "workers": state().scanner_gate.worker_count,
            "capacity": state().scanner_gate.capacity,
            "p4k_catalog_items": P4K_INVENTORY_CATALOG.item_count,
            "p4k_catalog_version": P4K_INVENTORY_CATALOG.version,
        },
        "item_catalog": catalog,
    }


@app.get("/api/me")
async def me(request: Request) -> dict[str, Any]:
    user = current_user_from_request(request, state().settings)
    if user is None:
        return {
            "authenticated": False,
            "discord_auth_enabled": discord_auth_configured(state().settings),
        }
    return {
        "authenticated": True,
        "id": user.id,
        "username": user.username,
        "display_name": user.display_name,
        "avatar_url": user.avatar_url,
        "roles": user.roles,
        "guild_permissions": user.guild_permissions,
        "can_manage_changes": user.can_manage_changes,
        "can_manage_admin": user.can_manage_admin,
    }


def _feedback_embed(
    *,
    report_type: str,
    details: str,
    expected_action: str,
    steps: str,
    recommendations: str,
    user: Any,
) -> dict[str, Any]:
    labels = {
        "issue": "Issue / Bug",
        "improvement": "Improvement recommendation",
        "feedback": "General feedback",
    }

    def field(name: str, value: str, *, inline: bool = False) -> dict[str, Any]:
        cleaned = value.strip() or "Not provided"
        if len(cleaned) > 1024:
            cleaned = cleaned[:1021].rstrip() + "..."
        return {"name": name, "value": cleaned, "inline": inline}

    embed: dict[str, Any] = {
        "title": "Star Citizen Companion Website Report",
        "color": 3587304,
        "fields": [
            field("Report type", labels.get(report_type, "Website report"), inline=True),
            field("Reported by", f"{user.display_name or user.username} (`{user.id}`)", inline=True),
            field("Issue / Feedback", details),
            field("Expected action or result", expected_action),
            field("Steps to reproduce", steps),
            field("Improvement recommendations", recommendations),
        ],
        "footer": {"text": "Submitted from sccompanion.org • Reply in this thread to add more information or images."},
    }
    if user.avatar_url:
        embed["author"] = {"name": user.display_name or user.username, "icon_url": user.avatar_url}
    return embed


def _provided_feedback_images(images: list[UploadFile]) -> list[UploadFile]:
    """Discard the empty multipart placeholder emitted for an unselected file input."""
    return [upload for upload in images if (upload.filename or "").strip()]


@app.post("/api/me/feedback")
async def submit_feedback(
    report_type: str = Form(...),
    title: str = Form(...),
    details: str = Form(...),
    expected_action: str = Form(default=""),
    steps: str = Form(default=""),
    recommendations: str = Form(default=""),
    images: list[UploadFile] = File(default=[]),
    user=Depends(require_user),
) -> dict[str, str]:
    settings = state().settings
    channel_id = settings.feedback_forum_channel_id
    if not channel_id or not settings.discord_token:
        raise HTTPException(status_code=503, detail="The Discord feedback forum is not configured.")
    if report_type not in {"issue", "improvement", "feedback"}:
        raise HTTPException(status_code=400, detail="Choose a valid report type.")
    title = " ".join(title.split())
    details = details.strip()
    if not title or not details:
        raise HTTPException(status_code=400, detail="A title and report details are required.")
    if len(title) > 100 or len(details) > 3000:
        raise HTTPException(status_code=400, detail="The report title or details are too long.")
    # Multipart forms may represent an unselected optional file input as an
    # UploadFile with an empty filename and application/octet-stream type.
    images = _provided_feedback_images(images)
    if len(images) > 4:
        raise HTTPException(status_code=400, detail="Attach no more than 4 images.")

    attachments: list[tuple[str, bytes, str]] = []
    total_bytes = 0
    allowed_types = {"image/png", "image/jpeg", "image/webp", "image/gif"}
    for index, upload in enumerate(images):
        content_type = (upload.content_type or "").lower()
        if content_type not in allowed_types:
            raise HTTPException(status_code=400, detail="Only PNG, JPG, WebP, and GIF images are supported.")
        data = await upload.read(8 * 1024 * 1024 + 1)
        if len(data) > 8 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Each image must be 8 MB or smaller.")
        total_bytes += len(data)
        if total_bytes > 20 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Combined image size must be 20 MB or smaller.")
        try:
            from PIL import Image
            Image.open(BytesIO(data)).verify()
        except Exception as error:
            raise HTTPException(status_code=400, detail="One of the attachments is not a valid image.") from error
        filename = re.sub(r"[^A-Za-z0-9._-]+", "-", Path(upload.filename or f"image-{index + 1}").name).strip(".-")
        attachments.append((filename or f"image-{index + 1}.png", data, content_type))

    message: dict[str, Any] = {
        "embeds": [_feedback_embed(
            report_type=report_type,
            details=details,
            expected_action=expected_action,
            steps=steps,
            recommendations=recommendations,
            user=user,
        )],
        "allowed_mentions": {"parse": []},
    }
    if attachments:
        message["attachments"] = [
            {"id": index, "filename": filename}
            for index, (filename, _data, _content_type) in enumerate(attachments)
        ]
    payload = {"name": title, "message": message}
    form = aiohttp.FormData()
    form.add_field("payload_json", json.dumps(payload), content_type="application/json")
    for index, (filename, data, content_type) in enumerate(attachments):
        form.add_field(f"files[{index}]", data, filename=filename, content_type=content_type)

    try:
        timeout = aiohttp.ClientTimeout(total=max(30, settings.http_timeout_seconds))
        async with aiohttp.ClientSession(timeout=timeout) as session:
            async with session.post(
                f"https://discord.com/api/v10/channels/{channel_id}/threads",
                headers={"Authorization": f"Bot {settings.discord_token}"},
                data=form,
            ) as response:
                response_payload = await response.json(content_type=None)
                if response.status >= 400:
                    logging.error("Discord feedback forum returned status %s", response.status)
                    raise HTTPException(status_code=502, detail="Discord could not create the feedback post.")
    except HTTPException:
        raise
    except (aiohttp.ClientError, asyncio.TimeoutError, ValueError) as error:
        logging.exception("Could not create Discord feedback forum post")
        raise HTTPException(status_code=502, detail="The feedback post could not reach Discord.") from error

    thread_id = str(response_payload.get("id") or "")
    if not thread_id:
        raise HTTPException(status_code=502, detail="Discord created the post but did not confirm the ticket.")
    return {"status": "submitted"}


@app.post("/api/activity", status_code=204)
async def website_activity() -> None:
    return None


def _begin_discord_login(settings: Settings) -> RedirectResponse:
    state_token = oauth_state()
    response = RedirectResponse(build_discord_authorize_url(settings, state_token), status_code=303)
    response.set_cookie(
        OAUTH_STATE_COOKIE_NAME,
        state_token,
        httponly=True,
        secure=settings.discord_redirect_uri.startswith("https://"),
        samesite="lax",
        max_age=300,
    )
    return response


@app.get("/auth/discord/login", response_model=None)
async def discord_login() -> RedirectResponse | HTMLResponse:
    settings = state().settings
    if not discord_auth_configured(settings):
        raise HTTPException(status_code=503, detail="Discord OAuth is not configured.")
    if not human_verification_configured(settings):
        return _begin_discord_login(settings)
    site_key = html.escape(settings.turnstile_site_key, quote=True)
    return HTMLResponse(f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Human verification</title><script src="https://challenges.cloudflare.com/turnstile/v0/api.js" async defer></script>
<style>body{{margin:0;background:#07131d;color:#e7f7ff;font:16px system-ui;display:grid;min-height:100vh;place-items:center}}
main{{width:min(420px,calc(100% - 40px));padding:28px;border:1px solid #247493;border-radius:12px;background:#0d2130}}
h1{{margin-top:0}}button{{margin-top:20px;padding:11px 18px;background:#26a9df;color:#031018;border:0;border-radius:6px;font-weight:700;cursor:pointer}}</style>
</head><body><main><h1>Verify you are human</h1><p>Complete this check before joining or linking Discord.</p>
<form method="post" action="/auth/discord/login"><div class="cf-turnstile" data-sitekey="{site_key}" data-theme="dark"></div>
<button type="submit">Continue with Discord</button></form></main></body></html>""")


@app.post("/auth/discord/login")
async def discord_login_verified(
    request: Request,
    turnstile_response: str = Form(alias="cf-turnstile-response"),
) -> RedirectResponse:
    settings = state().settings
    if not discord_auth_configured(settings):
        raise HTTPException(status_code=503, detail="Discord OAuth is not configured.")
    if not human_verification_configured(settings):
        raise HTTPException(status_code=503, detail="Human verification is not configured.")
    client_ip = request.client.host if request.client else None
    if not await verify_human(settings, turnstile_response, client_ip):
        raise HTTPException(status_code=400, detail="Human verification failed. Please go back and try again.")
    return _begin_discord_login(settings)


@app.get("/auth/discord/callback")
async def discord_callback(
    request: Request,
    code: str,
    oauth_state_value: str = Query(alias="state"),
) -> RedirectResponse:
    settings = state().settings
    expected_state = request.cookies.get(OAUTH_STATE_COOKIE_NAME)
    if not expected_state or expected_state != oauth_state_value:
        raise HTTPException(status_code=400, detail="Discord login state did not match.")
    token_payload = await exchange_discord_code(settings, code)
    user = await fetch_web_user(settings, str(token_payload.get("access_token")))
    secret = session_secret(settings)
    if not secret:
        raise HTTPException(status_code=503, detail="WEB_SESSION_SECRET or DISCORD_CLIENT_SECRET is required.")
    response = RedirectResponse("/")
    response.set_cookie(
        SESSION_COOKIE_NAME,
        encode_session(user, secret),
        httponly=True,
        secure=settings.discord_redirect_uri.startswith("https://"),
        samesite="lax",
        max_age=7 * 24 * 60 * 60,
    )
    response.delete_cookie(OAUTH_STATE_COOKIE_NAME)
    return response


@app.post("/auth/logout")
async def logout() -> RedirectResponse:
    response = RedirectResponse("/", status_code=303)
    response.delete_cookie(SESSION_COOKIE_NAME)
    return response


@app.get("/api/commands")
async def commands() -> dict[str, str]:
    return {"markdown": COMMANDS_PATH.read_text(encoding="utf-8")}


@app.get("/api/updates")
async def citizen_updates() -> dict[str, Any]:
    return await state().updates.get_updates()


@app.get("/api/lookup")
async def lookup(query: str) -> dict[str, Any]:
    result = await state().sources.lookup(query)
    if result is None:
        not_found(f"No result found for {query}.")
    return encode(result)


@app.get("/api/autocomplete/ships")
async def autocomplete_ships(query: str = "") -> list[str]:
    return await state().sources.autocomplete_ships(query)


@app.get("/api/ships/facets")
async def ship_facets() -> dict[str, list[str]]:
    return await state().sources.ship_facets()


@app.get("/api/ships/warbonds")
async def active_warbonds(refresh: bool = False) -> dict[str, Any]:
    return await state().warbonds.active(force_refresh=refresh)


@app.get("/api/ships")
async def ships(
    query: str | None = None,
    manufacturer: str | None = None,
    vehicle_type: str | None = None,
    size: str | None = None,
    role: str | None = None,
    status: str | None = None,
    sort_by: str = "name",
    min_cargo: float | None = Query(default=None, ge=0),
    max_cargo: float | None = Query(default=None, ge=0),
    limit: int = Query(default=24, ge=1, le=100),
    page: int = Query(default=1, ge=1),
) -> list[dict[str, Any]]:
    if min_cargo is not None and max_cargo is not None and min_cargo > max_cargo:
        raise HTTPException(status_code=422, detail="Minimum cargo cannot be greater than maximum cargo.")
    results = await state().sources.search_ships(
            query,
            manufacturer,
            vehicle_type,
            size,
            role,
            status,
            min_cargo,
            max_cargo,
            1000,
            1,
        )
    if sort_by == "cargo":
        results = sorted(results, key=lambda ship: float(ship.cargo_capacity or 0), reverse=True)
    elif sort_by == "manufacturer":
        results = sorted(results, key=lambda ship: ((ship.manufacturer or "").lower(), ship.name.lower()))
    elif sort_by == "size":
        results = sorted(results, key=lambda ship: ((ship.size or "").lower(), ship.name.lower()))
    else:
        results = sorted(results, key=lambda ship: ship.name.lower())
    start = max(0, (page - 1) * limit)
    return encode(results[start : start + limit])


@app.get("/api/ships/{name}")
async def ship(name: str) -> dict[str, Any]:
    result = await state().sources.lookup_ship(name)
    if result is None:
        not_found(f"No ship or vehicle found for {name}.")
    return encode(result)


@app.get("/api/me/ships")
async def my_ships(user=Depends(require_user)) -> list[dict[str, Any]]:
    ships = await state().cache.user_ships(user.id)
    repaired = False
    for ship in ships:
        display_name = _ship_display_name(str(ship.get("name") or ""))
        display_loaner_for = _ship_display_name(str(ship.get("loaner_for") or "")) if ship.get("loaner_for") else None
        cleaned_notes = _clean_redundant_loaner_note(
            ship.get("notes"),
            display_loaner_for,
        ) if ship.get("ownership_type") == "loaner" else ship.get("notes")
        if display_name and (
            display_name != ship.get("name")
            or display_loaner_for != ship.get("loaner_for")
            or cleaned_notes != ship.get("notes")
        ):
            await state().cache.save_user_ship(
                user.id,
                display_name,
                str(ship.get("ownership_type") or "in_game"),
                ship.get("manufacturer"),
                ship.get("role"),
                ship.get("source_name"),
                ship.get("source_url"),
                ship.get("image_url"),
                cleaned_notes,
                display_loaner_for,
            )
            if display_name != ship.get("name"):
                await state().cache.delete_user_ship(user.id, str(ship.get("name")))
            ship["name"] = display_name
            ship["loaner_for"] = display_loaner_for
            ship["notes"] = cleaned_notes
            repaired = True
        if not _ship_image_needs_refresh(ship.get("image_url")) and ship.get("manufacturer") and _has_ship_basic_info(ship.get("role")):
            if ship.get("ownership_type") == "pledged":
                repaired = await _sync_auto_loaners(user.id, str(ship.get("name") or ""), "pledged") or repaired
            continue
        detail = await state().sources.lookup_ship(str(ship.get("name") or ""))
        if ship.get("ownership_type") == "pledged":
            repaired = await _sync_auto_loaners(
                user.id,
                str(ship.get("name") or ""),
                "pledged",
                detail.status if detail else None,
            ) or repaired
        if detail is None:
            continue
        await state().cache.save_user_ship(
            user.id,
            _ship_display_name(detail.name),
            str(ship.get("ownership_type") or "in_game"),
            detail.manufacturer or ship.get("manufacturer"),
            _ship_basic_info(detail) or ship.get("role"),
            detail.source_name or ship.get("source_name"),
            detail.source_url or ship.get("source_url"),
            detail.image_url,
            ship.get("notes"),
            _ship_display_name(str(ship.get("loaner_for") or "")) if ship.get("loaner_for") else None,
        )
        repaired = True
    return await state().cache.user_ships(user.id) if repaired else ships


def _ship_image_needs_refresh(image_url: object) -> bool:
    value = str(image_url or "").strip().lower()
    return not value or "/thumb/" in value or "store_small" in value


@app.put("/api/me/ships")
async def save_my_ship(request: ShipOwnershipRequest, user=Depends(require_user)) -> dict[str, str]:
    ownership_type = request.ownership_type.strip().lower()
    if ownership_type not in {"pledged", "loaner", "in_game"}:
        raise HTTPException(status_code=422, detail="Ship ownership type must be pledged, loaner, or in_game.")
    ship_name = request.name.strip()
    display_name = _ship_display_name(ship_name)
    quantity = request.quantity
    if request.increment:
        existing = next(
            (ship for ship in await state().cache.user_ships(user.id) if _normalize_text(ship.get("name")) == _normalize_text(display_name)),
            None,
        )
        quantity = int(existing.get("quantity") or 1) + 1 if existing else 1
    await state().cache.save_user_ship(
        user.id,
        display_name,
        ownership_type,
        request.manufacturer.strip() if request.manufacturer else None,
        _ship_basic_info_from_values(
            request.role,
            request.vehicle_type,
            request.size,
            request.status,
            request.cargo_capacity,
        ),
        request.source_name.strip() if request.source_name else None,
        str(request.source_url).strip() if request.source_url else None,
        str(request.image_url).strip() if request.image_url else None,
        request.notes.strip() if request.notes else None,
        None,
        quantity,
    )
    await _sync_auto_loaners(user.id, display_name, ownership_type, request.status)
    return {"status": "saved"}


@app.post("/api/me/ships/import/rsi")
async def import_rsi_pledges(request: RsiPledgeImportRequest, user=Depends(require_user)) -> dict[str, Any]:
    imported: list[str] = []
    skipped: list[str] = []
    removed: list[str] = []
    candidates: set[str] = {
        cleaned
        for value in request.candidates[:500]
        if (cleaned := _clean_rsi_pledge_ship_name(value))
    }
    for page in request.pages:
        candidates.update(_extract_rsi_pledge_ship_names(page))
    if not candidates:
        raise HTTPException(status_code=400, detail="No ship or vehicle candidates were supplied.")
    for candidate in sorted(candidates, key=str.lower):
        detail = await _resolve_imported_ship(candidate)
        if detail is None:
            skipped.append(candidate)
            continue
        display_name = _ship_display_name(detail.name)
        await state().cache.save_user_ship(
            user.id,
            display_name,
            "pledged",
            detail.manufacturer,
            _ship_basic_info(detail),
            detail.source_name,
            detail.source_url,
            detail.image_url,
            None,
            None,
        )
        await _sync_auto_loaners(user.id, display_name, "pledged", detail.status)
        imported.append(display_name)
    # Candidate-based imports come from the extension's complete paginated scan.
    # Saved HTML uploads remain additive because users may upload only some pages.
    if request.candidates:
        protected_names = _rsi_import_protected_names(candidates, imported)
        existing_ships = await state().cache.user_ships(user.id)
        for ship in existing_ships:
            if ship.get("ownership_type") != "pledged":
                continue
            ship_name = str(ship.get("name") or "").strip()
            if not ship_name or _normalize_text(ship_name) in protected_names:
                continue
            await state().cache.delete_user_ship(user.id, ship_name)
            await state().cache.delete_user_loaners_for_ship(user.id, ship_name)
            removed.append(ship_name)
    return {
        "status": "imported",
        "candidates": sorted(candidates, key=str.lower),
        "imported": sorted(set(imported), key=str.lower),
        "skipped": sorted(set(skipped), key=str.lower),
        "removed": sorted(set(removed), key=str.lower),
    }


def _rsi_import_protected_names(candidates: set[str], imported: list[str]) -> set[str]:
    protected = {_normalize_text(name) for name in [*candidates, *imported]}
    for candidate in candidates:
        protected.update(_normalize_text(name) for name in _rsi_import_lookup_candidates(candidate))
    return protected


async def _resolve_imported_ship(candidate: str) -> Any:
    for name in _rsi_import_lookup_candidates(candidate):
        detail = await state().sources.lookup_ship(name)
        if detail is not None:
            return detail
    for name in _rsi_import_lookup_candidates(candidate):
        results = await state().sources.search_ships(query=name, limit=5)
        if results:
            normalized = _normalize_text(name)
            exact = next(
                (
                    ship
                    for ship in results
                    if _normalize_text(ship.name) == normalized
                    or _normalize_text(_ship_display_name(ship.name)) == normalized
                ),
                None,
            )
            return exact or results[0]
    return None


@app.delete("/api/me/ships/{ship_name}")
async def delete_my_ship(ship_name: str, user=Depends(require_user)) -> dict[str, str]:
    await state().cache.delete_user_ship(user.id, ship_name)
    await state().cache.delete_user_loaners_for_ship(user.id, ship_name)
    return {"status": "removed"}


async def _sync_auto_loaners(
    user_id: int,
    ship_name: str,
    ownership_type: str,
    status_hint: str | None = None,
) -> bool:
    loaner_names = SHIP_LOANERS.get(_normalize_text(ship_name), [])
    if not loaner_names:
        return False
    if ownership_type != "pledged":
        await state().cache.delete_user_loaners_for_ship(user_id, ship_name)
        return True
    detail = None
    if not _ship_is_in_concept(status_hint):
        detail = await state().sources.lookup_ship(ship_name)
        if not _ship_is_in_concept(detail.status if detail else None):
            await state().cache.delete_user_loaners_for_ship(user_id, ship_name)
            return True
    existing_ships = await state().cache.user_ships(user_id)
    existing_by_name = {
        _normalize_text(existing.get("name")): existing
        for existing in existing_ships
    }
    changed = False
    for loaner_name in loaner_names:
        loaner_display_name = _ship_display_name(loaner_name)
        existing = existing_by_name.get(_normalize_text(loaner_name)) or existing_by_name.get(_normalize_text(loaner_display_name))
        if existing and existing.get("loaner_for") != ship_name:
            continue
        loaner = await state().sources.lookup_ship(loaner_name)
        await state().cache.save_user_ship(
            user_id,
            _ship_display_name(loaner.name if loaner else loaner_display_name),
            "loaner",
            loaner.manufacturer if loaner else None,
            _ship_basic_info(loaner) if loaner else None,
            loaner.source_name if loaner else "RSI Loaner Ship Matrix",
            loaner.source_url if loaner else "https://support.robertsspaceindustries.com/hc/en-us/articles/360003093114-Loaner-Ship-Matrix",
            loaner.image_url if loaner else None,
            None,
            ship_name,
        )
        changed = True
    return changed


def _normalize_text(value: object) -> str:
    return " ".join(str(value or "").lower().replace("-", " ").split())


def _clean_redundant_loaner_note(notes: object, loaner_for: str | None) -> str | None:
    value = str(notes or "").strip()
    if not value:
        return None
    if loaner_for and _normalize_text(value) == _normalize_text(f"Loaner for {loaner_for}"):
        return None
    return value


def _extract_rsi_pledge_ship_names(page_html: str) -> set[str]:
    candidates: set[str] = set()
    candidates.update(_extract_rsi_typed_item_ship_names(page_html))
    candidates.update(_extract_rsi_pledge_ship_names_from_links(page_html))
    candidates.update(_extract_rsi_pledge_ship_names_from_json(page_html))
    text = html.unescape(re.sub(r"<[^>]+>", "\n", page_html))
    text = re.sub(r"\s+", " ", text)
    candidates.update(_extract_rsi_pledge_ship_names_from_blocks(text))
    patterns = [
        r"(?:Contains|Also Contains)\s*:?\s+([^$<>]{2,120}?)(?=\s+(?:Also Contains|Standalone Ship|Package|Serial|Insurance|Starting Money|Hangar|Downloadable|Contains|$))",
        r"(?:Standalone Ship|Game Package|Package)\s*[-:]\s*([^$<>]{2,100}?)(?=\s+(?:Attributed|Created|Serial|Insurance|Contains|$))",
        r"Ship\s*[:\-]\s*([^$<>]{2,100}?)(?=\s+(?:Serial|Insurance|Contains|$))",
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, text, flags=re.IGNORECASE):
            for name in re.split(r"\s+(?:and|&|\+)\s+", match.group(1)):
                cleaned = _clean_rsi_pledge_ship_name(name)
                if cleaned:
                    candidates.add(cleaned)
    return candidates


def _extract_rsi_typed_item_ship_names(page_html: str) -> set[str]:
    candidates: set[str] = set()
    item_starts = [
        match.start()
        for match in re.finditer(r'''<[^>]+class=["'][^"']*\bitem\b[^"']*["'][^>]*>''', page_html, flags=re.IGNORECASE)
    ]
    for index, start in enumerate(item_starts):
        end = min(item_starts[index + 1] if index + 1 < len(item_starts) else len(page_html), start + 2400)
        block = page_html[start:end]
        kind_match = re.search(r'''class=["'][^"']*\bkind\b[^"']*["'][^>]*>([\s\S]{0,240}?)</[^>]+>''', block, flags=re.IGNORECASE)
        title_match = re.search(r'''class=["'][^"']*\btitle\b[^"']*["'][^>]*>([\s\S]{0,240}?)</[^>]+>''', block, flags=re.IGNORECASE)
        if not kind_match or not title_match:
            continue
        kind_text = html.unescape(re.sub(r"<[^>]+>", " ", kind_match.group(1)))
        if not re.search(r"\b(?:ship|vehicle)\b", kind_text, flags=re.IGNORECASE):
            continue
        title_text = html.unescape(re.sub(r"<[^>]+>", " ", title_match.group(1)))
        cleaned = _clean_rsi_pledge_ship_name(title_text)
        if cleaned:
            candidates.add(cleaned)
    return candidates


def _extract_rsi_pledge_ship_names_from_links(page_html: str) -> set[str]:
    candidates: set[str] = set()
    for match in re.finditer(r"/pledge/ships/[^\"'<> ]+/([^\"'<>?#]+)", page_html, flags=re.IGNORECASE):
        slug = html.unescape(match.group(1))
        cleaned = _clean_rsi_pledge_ship_name(slug.replace("-", " "))
        if cleaned:
            candidates.add(cleaned)
    return candidates


def _extract_rsi_pledge_ship_names_from_blocks(text: str) -> set[str]:
    candidates: set[str] = set()
    block_pattern = (
        r"(?P<kind>Standalone Ship|Game Package|Package)\s+"
        r"(?P<body>.{0,900}?)(?=\s+(?:Standalone Ship|Game Package|Package|Upgrade|Add-Ons|Paints|$))"
    )
    for block in re.finditer(block_pattern, text, flags=re.IGNORECASE):
        body = block.group("body")
        for pattern in (
            r"(?:Contains|Also Contains)\s*:?\s+([^$<>]{2,140}?)(?=\s+(?:Also Contains|Attributed|Created|Serial|Insurance|Starting Money|Hangar|Downloadable|Contains|$))",
            r"^\s*[-:]?\s*([^$<>]{2,120}?)(?=\s+(?:Attributed|Created|Serial|Insurance|Contains|$))",
        ):
            for match in re.finditer(pattern, body, flags=re.IGNORECASE):
                cleaned = _clean_rsi_pledge_ship_name(match.group(1))
                if cleaned:
                    candidates.add(cleaned)
    return candidates


def _extract_rsi_pledge_ship_names_from_json(page_html: str) -> set[str]:
    candidates: set[str] = set()
    for key in ("name", "title", "label"):
        pattern = rf'"{key}"\s*:\s*"([^"]{{2,120}})"'
        for match in re.finditer(pattern, page_html, flags=re.IGNORECASE):
            raw_value = match.group(1).encode("utf-8").decode("unicode_escape", errors="ignore")
            # RSI pages contain JSON labels for paints, equipment, flair, currencies,
            # navigation, and recommendations. Only pledge titles that explicitly
            # identify a ship or ship-bearing package belong in the hangar import.
            if not re.match(r"^\s*(?:Standalone Ship|Game Package|Package)\s*(?:[-:]|\s)", raw_value, flags=re.IGNORECASE):
                continue
            cleaned = _clean_rsi_pledge_ship_name(raw_value)
            if cleaned:
                candidates.add(cleaned)
    return candidates


def _clean_rsi_pledge_ship_name(name: str) -> str | None:
    value = " ".join(name.split())
    value = re.sub(r"^(?:Standalone Ship|Package|Upgrade|Ship)\s*[-:]\s*", "", value, flags=re.IGNORECASE)
    value = re.sub(r"\b(?:with Lifetime Insurance|with Lifetime|Lifetime Insurance|Best In Show|BIS|ILW|IAE|Warbond|Edition|Paint|Poster|Model|Serial|LTI)\b.*$", "", value, flags=re.IGNORECASE).strip(" -:,.")
    if not value or len(value) < 2:
        return None
    if len(value) > 72 or len(value.split()) > 10:
        return None
    blocked_words = {"insurance", "hangar", "poster", "paint", "skin", "flair", "manual", "downloadable"}
    blocked_terms = {
        "add on",
        "add-on",
        "attributed",
        "canadian dollar",
        "created",
        "download",
        "downloadable",
        "english",
        "figurine",
        "flair",
        "gift card",
        "gold livery",
        "hangar",
        "name reservation",
        "paint",
        "plushie",
        "poster",
        "pound sterling",
        "referral reward",
        "self land",
        "skin",
        "store",
        "subscriber",
        "support",
        "upgrade",
        "united states dollar",
    }
    normalized = _normalize_text(value)
    if re.search(r"\bto\b.*\b(?:year|insurance)\b|\b\d+\s*year\b", normalized):
        return None
    if normalized in blocked_words or any(term in normalized for term in blocked_terms):
        return None
    return value


def _rsi_import_lookup_candidates(value: str) -> list[str]:
    cleaned = _clean_rsi_pledge_ship_name(value) or value
    candidates = [cleaned, _ship_display_name(cleaned)]
    normalized = _normalize_text(cleaned)
    package_markers = [
        "starter pack",
        "starter package",
        "game package",
        "pack",
        "package",
    ]
    simplified = cleaned
    for marker in package_markers:
        simplified = re.sub(rf"\b{re.escape(marker)}\b", "", simplified, flags=re.IGNORECASE)
    simplified = " ".join(simplified.split()).strip(" -:,.")
    if simplified:
        candidates.extend([simplified, _ship_display_name(simplified)])
    if " - " in cleaned:
        candidates.append(cleaned.split(" - ")[-1].strip())
    result: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        if not candidate:
            continue
        key = _normalize_text(candidate)
        if key in seen:
            continue
        seen.add(key)
        result.append(candidate)
    return result


def _ship_is_in_concept(status: object) -> bool:
    return _normalize_text(status) == "in concept"


def _ship_basic_info(ship: Any) -> str | None:
    role = " / ".join(str(value) for value in [getattr(ship, "career", None), getattr(ship, "role", None)] if value)
    return _ship_basic_info_from_values(
        role,
        getattr(ship, "vehicle_type", None),
        getattr(ship, "size", None),
        getattr(ship, "status", None),
        getattr(ship, "cargo_capacity", None),
    )


def _ship_basic_info_from_values(
    role: object,
    vehicle_type: object,
    size: object,
    status: object,
    cargo_capacity: object,
) -> str | None:
    parts = [role, vehicle_type, size, status]
    if cargo_capacity is not None and cargo_capacity != "":
        parts.append(f"{cargo_capacity} SCU")
    return " | ".join(str(part) for part in parts if part) or None


def _has_ship_basic_info(value: object) -> bool:
    return " | " in str(value or "")


def _ship_display_name(name: str) -> str:
    value = " ".join(str(name or "").split())
    for prefix in SHIP_DISPLAY_PREFIXES:
        if value.startswith(prefix):
            return value.removeprefix(prefix).strip()
    return value


@app.get("/api/commodities/{name}")
async def commodity(
    name: str,
    system: str | None = None,
    purchase_system: str | None = None,
    sell_system: str | None = None,
) -> dict[str, Any]:
    result = await state().sources.lookup_commodity(name, system, purchase_system, sell_system)
    if result is None:
        not_found(f"No commodity found for {name}.")
    return encode(result)


@app.get("/api/autocomplete/commodities")
async def autocomplete_commodities(query: str = "") -> list[str]:
    return await state().sources.autocomplete_commodities(query)


@app.get("/api/mining/{material}")
async def mining(material: str, system: str | None = None, planet: str | None = None) -> dict[str, Any]:
    terms = _mining_multi_search_terms(material)
    result = None
    if len(terms) == 1 and not _has_mining_multi_separator(material):
        result = await state().sources.lookup_mining_material(material, system, planet)
        if result is None:
            terms = _mining_space_search_terms(material)

    if len(terms) > 1:
        return await multi_mining_signature_payload(material, terms)

    if result is None:
        result = await state().sources.lookup_mining_material(material, system, planet)
    if result is None:
        not_found(f"No mining material found for {material}.")
    return encode(await apply_community_mining_locations(state().cache, result))


async def multi_mining_signature_payload(query: str, terms: list[str]) -> dict[str, Any]:
    results = []
    missing = []
    location_materials: dict[tuple[str, str], set[str]] = {}
    for term in terms:
        result = await state().sources.lookup_mining_material(term)
        if result is None:
            missing.append(term)
            continue
        results.append(
            {
                "term": term,
                "material": result.material_name,
                "signatures": _mining_term_signatures(result, term),
            }
        )
        groups = result.location_groups or []
        if groups:
            for group in groups:
                locations = [*group.lagrange_points, *group.planets, *group.moons, *group.points_of_interest]
                for location in locations:
                    location_materials.setdefault((group.system, location), set()).add(result.material_name)
        else:
            locations = [*result.lagrange_points, *result.planets, *result.moons, *result.points_of_interest]
            for location in locations:
                location_materials.setdefault(("", location), set()).add(result.material_name)

    shared_signatures = _shared_mining_signatures([result["signatures"] for result in results])
    ranked_locations = sorted(
        (
            {"system": system or None, "location": location, "materials": sorted(materials), "match_count": len(materials)}
            for (system, location), materials in location_materials.items()
        ),
        key=lambda item: (-item["match_count"], str(item["system"] or ""), item["location"]),
    )[:30]
    return {
        "result_type": "multi_mining_signatures",
        "material_name": "Mining Signature Match",
        "query": query,
        "materials": _unique_preserve_order([result["material"] for result in results]),
        "missing": missing,
        "rock_signatures": shared_signatures,
        "ranked_locations": ranked_locations,
        "source_name": "Star Citizen Wiki mining signatures",
    }


@app.post("/api/mining/community", dependencies=[Depends(require_change_admin)])
async def mining_community(payload: MiningCommunityRequest, request: Request) -> dict[str, str]:
    await add_community_mining_location(state().cache, payload.model_dump())
    await state().cache.add_audit_event(
        "Website Mining Location Added",
        {
            "User": _website_audit_user(current_user_from_request(request, state().settings)),
            "Material": payload.material,
            "System": payload.system,
            "Location Type": payload.location_type,
            "Location": payload.location,
            "Reported By": payload.reported_by,
        },
        "mining",
    )
    return {"status": "saved"}


@app.get("/api/me/refinery-orders")
async def my_refinery_orders(user=Depends(require_user)) -> list[dict[str, Any]]:
    return await state().cache.user_refinery_orders(user.id)


@app.post("/api/me/refinery-orders")
async def save_my_refinery_order(
    payload: RefineryOrderRequest,
    user=Depends(require_user),
) -> dict[str, Any]:
    values = payload.model_dump()
    order_id = await state().cache.save_user_refinery_order(user.id, values)
    return {"status": "saved", "id": order_id}


@app.put("/api/me/refinery-orders/{order_id}/status")
async def update_my_refinery_order_status(
    order_id: int,
    payload: RefineryOrderStatusRequest,
    user=Depends(require_user),
) -> dict[str, str]:
    status = payload.status.strip().lower()
    if status not in {"refining", "ready", "collected", "sold", "archived"}:
        raise HTTPException(status_code=422, detail="Unsupported refinery order status.")
    if not await state().cache.update_user_refinery_order_status(user.id, order_id, status):
        not_found("Refinery order not found.")
    return {"status": "updated"}


@app.delete("/api/me/refinery-orders/{order_id}")
async def delete_my_refinery_order(order_id: int, user=Depends(require_user)) -> dict[str, str]:
    if not await state().cache.delete_user_refinery_order(user.id, order_id):
        not_found("Refinery order not found.")
    return {"status": "deleted"}


@app.get("/api/autocomplete/mining-materials")
async def autocomplete_mining_materials(query: str = "") -> list[str]:
    return await state().sources.autocomplete_mining_materials(query)


@app.get("/api/blueprints")
async def blueprints(
    query: str | None = None,
    category: str | None = None,
    material: str | None = None,
    mission_type: str | None = None,
    contractor: str | None = None,
    location: str | None = None,
    limit: int = Query(default=12, ge=1, le=50),
    page: int = Query(default=1, ge=1),
) -> list[dict[str, Any]]:
    return encode(
        await state().sources.lookup_blueprints(
            query,
            category,
            material,
            mission_type,
            contractor,
            location,
            limit,
            page,
        )
    )


@app.get("/api/autocomplete/blueprints")
async def autocomplete_blueprints(query: str = "") -> list[str]:
    return await state().sources.autocomplete_blueprints(query)


@app.post("/api/blueprints/import/text")
async def import_blueprints_from_text(request: BlueprintTextImportRequest) -> dict[str, Any]:
    return {
        "ocr_available": True,
        "ocr_text": request.text,
        "matches": encode(await _match_blueprints_from_text(request.text)),
    }


@app.post("/api/blueprints/import/images")
async def import_blueprints_from_images(files: list[UploadFile] = File(...)) -> dict[str, Any]:
    ocr_text, ocr_error = await _ocr_blueprint_images(files)
    return {
        "ocr_available": ocr_error is None,
        "ocr_error": ocr_error,
        "ocr_text": ocr_text,
        "matches": encode(await _match_blueprints_from_text(ocr_text)) if ocr_text.strip() else [],
    }


@app.get("/api/me/blueprints")
async def my_blueprints(user=Depends(require_user)) -> list[dict[str, Any]]:
    return await state().cache.user_blueprints(user.id)


@app.put("/api/me/blueprints")
async def save_my_blueprint(request: BlueprintOwnershipRequest, user=Depends(require_user)) -> dict[str, str]:
    await state().cache.save_user_blueprint(
        user.id,
        request.name.strip(),
        request.category.strip() if request.category else None,
        request.source_name.strip() if request.source_name else None,
        str(request.source_url).strip() if request.source_url else None,
    )
    return {"status": "saved"}


@app.delete("/api/me/blueprints/{blueprint_name}")
async def delete_my_blueprint(blueprint_name: str, user=Depends(require_user)) -> dict[str, str]:
    await state().cache.delete_user_blueprint(user.id, blueprint_name)
    return {"status": "removed"}


@app.get("/api/me/inventory")
async def my_inventory(
    location: str | None = None,
    category: str | None = None,
    query: str | None = None,
    sort_by: str = "name",
    user=Depends(require_user),
) -> list[dict[str, Any]]:
    return await state().cache.user_inventory_items(
        user.id,
        location.strip() if location else None,
        category.strip() if category else None,
        query.strip() if query else None,
        sort_by,
    )


@app.get("/api/me/inventory/facets")
async def my_inventory_facets(user=Depends(require_user)) -> dict[str, list[str]]:
    return await state().cache.user_inventory_facets(user.id)


@app.get("/api/me/inventory/catalog")
async def inventory_catalog_suggestions(
    query: str = Query(default="", max_length=120),
    category: str = Query(min_length=1, max_length=40),
    limit: int = Query(default=12, ge=1, le=25),
    user=Depends(require_user),
) -> list[dict[str, Any]]:
    """Autocomplete manual entry from the local Wiki and Data.p4k indexes."""
    del user
    clean_query = " ".join(query.split())
    clean_category = " ".join(category.split())
    if not clean_category:
        raise HTTPException(status_code=422, detail="Select a category before searching the game catalog.")
    if len(clean_query) < 2:
        return []
    # Manual autocomplete must remain instant even while the Wiki catalog is
    # syncing or unavailable. Data.p4k is bundled and queried entirely in
    # memory, so typing never depends on a remote request.
    p4k_results = P4K_INVENTORY_CATALOG.lookup(clean_query, limit * 2)
    merged: dict[str, Any] = {}
    for result in p4k_results:
        key = _normalize_text(getattr(result, "name", ""))
        if key and key not in merged:
            merged[key] = result
    query_norm = _normalize_text(clean_query)

    def rank(result: Any) -> tuple[float, float, float, str]:
        name = str(getattr(result, "name", ""))
        name_norm = _normalize_text(name)
        prefix = 1.0 if name_norm.startswith(query_norm) else 0.0
        query_words = re.findall(r"[a-z0-9]+", query_norm)
        name_words = re.findall(r"[a-z0-9]+", name_norm)
        token_coverage = sum(
            1 for word in query_words if any(candidate.startswith(word) for candidate in name_words)
        ) / max(len(query_words), 1)
        aliases = (name, *getattr(result, "catalog_aliases", ()))
        score = max(_inventory_match_confidence(clean_query, alias) for alias in aliases)
        return (-token_coverage, -prefix, -score, name.casefold())

    suggestions: list[dict[str, Any]] = []
    for result in sorted(merged.values(), key=rank):
        # Manual typing should narrow predictably by the visible item name;
        # OCR-oriented fuzzy aliases are useful to the scanner but make an
        # autocomplete query such as "bliz" return unrelated names.
        if query_norm not in _normalize_text(getattr(result, "name", "")):
            continue
        result_category = _inventory_manual_catalog_category(result)
        if not result_category or result_category.casefold() != clean_category.casefold():
            continue
        suggestions.append({
            "name": result.name,
            "category": result_category,
            "item_type": (
                _inventory_catalog_item_type(result.name, result_category)
                or getattr(result, "section", None)
            ),
            "item_size": getattr(result, "size", None),
            "source_name": getattr(result, "source_name", None),
        })
        if len(suggestions) >= limit:
            break
    return suggestions


@app.post("/api/me/inventory/import/text")
async def import_inventory_from_text(
    request: InventoryTextImportRequest,
    user=Depends(require_user),
) -> dict[str, Any]:
    del user
    if request.scanner_mode:
        if not (request.default_category or "").strip():
            raise HTTPException(
                status_code=422,
                detail="Select the matching in-game inventory category before scanning.",
            )
        scanner_lookups = await _inventory_scanner_lookups(
            request.text,
            request.exclude_words,
            category=request.default_category,
            item_type=request.default_item_type,
        )
        return {
            "ocr_available": True,
            "ocr_text": request.text,
            "items": await _match_inventory_scanner_text(
                request.text,
                request.default_location,
                request.default_category,
                request.min_score,
                request.exclude_words,
                scanner_lookups,
            ),
            "diagnostics": await _inventory_scanner_diagnostics(
                request.text,
                request.min_score,
                request.exclude_words,
                scanner_lookups,
            ),
        }
    return {
        "ocr_available": True,
        "ocr_text": request.text,
        "items": await _enrich_inventory_items(
            _inventory_items_from_text(
                request.text,
                request.default_location,
                request.default_category,
                first_match=request.scanner_mode,
            )
        ),
    }


@app.post("/api/me/inventory/import/images")
async def import_inventory_from_images(
    files: list[UploadFile] = File(...),
    default_location: str | None = None,
    default_category: str | None = None,
    default_item_type: str | None = None,
    scanner_mode: bool = False,
    live_scan: bool = False,
    title_box: str | None = None,
    min_score: float = Query(default=0.72, ge=0, le=1),
    exclude_words: str | None = None,
    user=Depends(require_user),
) -> dict[str, Any]:
    if scanner_mode and not (default_category or "").strip():
        raise HTTPException(
            status_code=422,
            detail="Select the matching in-game inventory category before scanning.",
        )
    if scanner_mode:
        async with state().scanner_gate.admit(user.id) as (scan_id, queue_ms):
            return await _import_inventory_scanner_images(
                files,
                default_location,
                default_category,
                default_item_type,
                live_scan,
                title_box,
                min_score,
                exclude_words,
                scan_id,
                queue_ms,
            )

    ocr_text, ocr_error = await _ocr_blueprint_images(files)
    return {
        "ocr_available": ocr_error is None,
        "ocr_error": ocr_error,
        "ocr_text": ocr_text,
        "items": await _enrich_inventory_items(
            _inventory_items_from_text(ocr_text, default_location, default_category)
        ) if ocr_text.strip() else [],
    }


async def _import_inventory_scanner_images(
    files: list[UploadFile],
    default_location: str | None,
    default_category: str | None,
    default_item_type: str | None,
    live_scan: bool,
    title_box: str | None,
    min_score: float,
    exclude_words: str | None,
    scan_id: str,
    queue_ms: int,
) -> dict[str, Any]:
    started_at = time.perf_counter()
    ocr_started_at = time.perf_counter()
    calibration: dict[str, Any] | None = None
    effective_min_score = max(min_score, 0.88) if live_scan else min_score
    if not live_scan:
        ocr_text, ocr_error = await _ocr_blueprint_images(files)
        ocr_ms = round((time.perf_counter() - ocr_started_at) * 1000)
        match_started_at = time.perf_counter()
        scanner_lookups = await _inventory_scanner_lookups(
            ocr_text,
            exclude_words,
            category=default_category,
            item_type=default_item_type,
        ) if ocr_text.strip() else {}
        items = await _match_inventory_scanner_text(
            ocr_text,
            default_location,
            default_category,
            effective_min_score,
            exclude_words,
            scanner_lookups,
        ) if ocr_text.strip() else []
        match_ms = round((time.perf_counter() - match_started_at) * 1000)
        attempted_titles: list[tuple[str, str]] = []
    else:
        image_data, ocr_error = await _read_inventory_scanner_image(files)
        ocr_text = ""
        scanner_lookups: dict[str, list[tuple[Any, float]]] = {}
        items: list[dict[str, Any]] = []
        attempted_titles = []
        best_review: tuple[float, str, str, dict[str, list[tuple[Any, float]]]] | None = None
        candidate_boxes = _inventory_title_boxes(title_box)
        candidate_texts = await asyncio.gather(*(
            asyncio.to_thread(_read_calibrated_inventory_title, image_data, candidate_box)
            for candidate_box in candidate_boxes
        )) if image_data else [""] * len(candidate_boxes)
        ocr_ms = round((time.perf_counter() - ocr_started_at) * 1000)
        plausible_titles: list[tuple[str, str]] = []
        seen_titles: set[str] = set()
        for candidate_box, candidate_text in zip(candidate_boxes, candidate_texts):
            if not _inventory_title_candidate_is_plausible(candidate_text):
                continue
            normalized_title = _normalize_text(candidate_text)
            if normalized_title in seen_titles:
                continue
            seen_titles.add(normalized_title)
            attempted_titles.append((candidate_text, candidate_box))
            plausible_titles.append((candidate_text, candidate_box))
        match_started_at = time.perf_counter()
        lookup_groups = await asyncio.gather(*(
            _inventory_scanner_lookups(
                candidate_text, exclude_words, candidate_limit=1,
                category=default_category, item_type=default_item_type,
            )
            for candidate_text, _ in plausible_titles
        ))
        for (candidate_text, candidate_box), candidate_lookups in zip(plausible_titles, lookup_groups):
            candidate_items = await _match_inventory_scanner_text(
                candidate_text,
                default_location,
                default_category,
                effective_min_score,
                exclude_words,
                candidate_lookups,
            )
            best_score = max(
                (
                    score
                    for matches in candidate_lookups.values()
                    for _, score in matches
                ),
                default=0.0,
            )
            if best_review is None or best_score > best_review[0]:
                best_review = (best_score, candidate_text, candidate_box, candidate_lookups)
            if candidate_items:
                ocr_text = candidate_text
                scanner_lookups = candidate_lookups
                items = candidate_items
                calibration = {"title_box": candidate_box, "fast_title": True}
                break
        if not items and best_review is not None:
            _, ocr_text, candidate_box, scanner_lookups = best_review
            calibration = {"title_box": candidate_box, "fast_title": False}
        elif calibration is None:
            calibration = {"title_box": None, "fast_title": False}
        match_ms = round((time.perf_counter() - match_started_at) * 1000)
    logging.info(
        "Inventory scanner scan_id=%s category=%r type=%r ocr=%r matches=%r attempts=%r queue_ms=%d ocr_ms=%d match_ms=%d",
        scan_id,
        default_category,
        default_item_type,
        " | ".join(ocr_text.splitlines())[:500],
        [item.get("name") for item in items],
        [(text[:120], box) for text, box in attempted_titles],
        queue_ms,
        ocr_ms,
        match_ms,
    )
    return {
        "scan_id": scan_id,
        "ocr_available": ocr_error is None,
        "ocr_error": ocr_error,
        "ocr_text": ocr_text,
        "items": items,
        "calibration": calibration,
        "diagnostics": await _inventory_scanner_diagnostics(
            ocr_text,
            effective_min_score,
            exclude_words,
            scanner_lookups,
        ) if ocr_text.strip() else {"candidates": [], "rejected_lines": []},
        "performance": {
            "queue_ms": queue_ms,
            "ocr_ms": ocr_ms,
            "match_ms": match_ms,
            "server_ms": round((time.perf_counter() - started_at) * 1000),
        },
    }


@app.post("/api/me/inventory")
async def add_my_inventory_item(request: InventoryItemRequest, user=Depends(require_user)) -> dict[str, Any]:
    await state().cache.merge_user_inventory_duplicates(user.id)
    item_id = await state().cache.save_user_inventory_item(
        user.id,
        request.name.strip(),
        request.category.strip() if request.category else None,
        request.location.strip(),
        request.quantity,
        request.quality,
        request.item_type.strip() if request.item_type else None,
        request.item_size.strip() if request.item_size else None,
        request.volume_scu,
        request.notes.strip() if request.notes else None,
    )
    return {"status": "saved", "id": item_id}


@app.post("/api/me/inventory/merge-duplicates")
async def merge_my_inventory_duplicates(user=Depends(require_user)) -> dict[str, Any]:
    removed = await state().cache.merge_user_inventory_duplicates(user.id)
    return {"status": "merged", "removed": removed}


@app.post("/api/me/inventory/clear")
async def clear_my_inventory(request: InventoryClearRequest, user=Depends(require_user)) -> dict[str, Any]:
    location = request.location.strip() if request.location else None
    removed = await state().cache.clear_user_inventory_items(user.id, location or None)
    return {"status": "cleared", "removed": removed, "location": location}


@app.get("/api/me/inventory/export")
async def export_my_inventory(
    location: str | None = None,
    category: list[str] = Query(default=[]),
    query: str | None = None,
    sort_by: str = "location",
    selling: bool = False,
    user=Depends(require_user),
) -> Response:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    items = await state().cache.user_inventory_items(
        user.id,
        location.strip() if location else None,
        None,
        query.strip() if query else None,
        sort_by,
    )
    selected_categories = {value.strip().casefold() for value in category if value.strip()}
    if selected_categories:
        items = [item for item in items if str(item.get("category") or "").strip().casefold() in selected_categories]

    price_comparisons: dict[str, dict[str, float]] = {}
    if selling and items:
        try:
            price_comparisons = await state().sources.inventory_sell_price_comparison([str(item["name"]) for item in items])
        except Exception:
            logging.exception("UEX inventory selling prices were unavailable during export")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Station Inventory"
    headers = ["Location", "Category", "Item Type", "Quantity", "Name", "Size", "Quality", "Volume SCU", "Notes"]
    if selling:
        headers.extend([
            "Average UEX Terminal Sell Price (aUEC)",
            "Average UEX Player Seller Price (aUEC)",
            "Price Source",
            "Estimated Sell Total (aUEC)",
        ])
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for item in items:
        row = [
                item["location"],
                item["category"] or "",
                item["item_type"] or "",
                item["quantity"],
                item["name"],
                item["item_size"] or "",
                item["quality"] if item["quality"] is not None else "",
                item["volume_scu"] if item["volume_scu"] is not None else "",
                "",
            ]
        if selling:
            normalized_name = " ".join(re.sub(r"[^a-z0-9]+", " ", str(item["name"]).casefold()).split())
            comparison = price_comparisons.get(normalized_name, {})
            terminal_price = comparison.get("terminal_average")
            player_price = comparison.get("player_average")
            selected_price = player_price if player_price is not None else terminal_price
            row.extend([
                terminal_price if terminal_price is not None else "",
                player_price if player_price is not None else "",
                "UEX Player Marketplace" if player_price is not None else ("UEX Terminal Buyback" if terminal_price is not None else ""),
                selected_price * float(item["quantity"]) if selected_price is not None else "",
            ])
        sheet.append(row)
    if selling:
        for row_number in range(2, sheet.max_row + 1):
            for column_number in (10, 11, 13):
                sheet.cell(row=row_number, column=column_number).number_format = '#,##0.00'
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 48)
    output = BytesIO()
    workbook.save(output)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="station-inventory.xlsx"'},
    )


@app.put("/api/me/inventory/{item_id}")
async def update_my_inventory_item(
    item_id: int,
    request: InventoryItemRequest,
    user=Depends(require_user),
) -> dict[str, str]:
    updated = await state().cache.update_user_inventory_item(
        user.id,
        item_id,
        request.name.strip(),
        request.category.strip() if request.category else None,
        request.location.strip(),
        request.quantity,
        request.quality,
        request.item_type.strip() if request.item_type else None,
        request.item_size.strip() if request.item_size else None,
        request.volume_scu,
        request.notes.strip() if request.notes else None,
    )
    if not updated:
        not_found("Inventory item not found.")
    return {"status": "updated"}


@app.post("/api/me/inventory/{item_id}/transfer")
async def transfer_my_inventory_item(
    item_id: int,
    request: InventoryTransferRequest,
    user=Depends(require_user),
) -> dict[str, str]:
    updated = await state().cache.transfer_user_inventory_item(user.id, item_id, request.location.strip())
    if not updated:
        not_found("Inventory item not found.")
    return {"status": "transferred"}


@app.delete("/api/me/inventory/{item_id}")
async def delete_my_inventory_item(item_id: int, user=Depends(require_user)) -> dict[str, str]:
    deleted = await state().cache.delete_user_inventory_item(user.id, item_id)
    if not deleted:
        not_found("Inventory item not found.")
    return {"status": "removed"}


async def _ocr_blueprint_images(files: list[UploadFile]) -> tuple[str, str | None]:
    try:
        from PIL import Image
    except Exception:
        return "", "Image OCR needs Pillow installed on the server."

    texts: list[str] = []
    for file in files[:8]:
        data = await file.read(state().settings.max_upload_bytes + 1)
        if len(data) > state().settings.max_upload_bytes:
            raise HTTPException(status_code=413, detail="Uploaded image is too large.")
        if not data:
            continue
        try:
            Image.open(BytesIO(data)).verify()
            texts.append(await asyncio.to_thread(_read_image_text, data))
        except Exception as exc:
            return "\n".join(texts), f"Could not read {file.filename or 'image'}: {exc}"
    return "\n".join(texts), None


async def _ocr_inventory_title_images(
    files: list[UploadFile],
    title_box: str | None,
) -> tuple[str, str | None, str | None, bool]:
    for file in files[:1]:
        data = await file.read(state().settings.max_upload_bytes + 1)
        if len(data) > state().settings.max_upload_bytes:
            raise HTTPException(status_code=413, detail="Uploaded image is too large.")
        if not data:
            continue
        try:
            text, calibrated_box, used_fast = await asyncio.to_thread(
                _read_inventory_title,
                data,
                title_box,
            )
            return text, None, calibrated_box, used_fast
        except Exception as exc:
            return "", f"Could not read {file.filename or 'image'}: {exc}", title_box, False
    return "", None, title_box, False


async def _ocr_inventory_title_candidates(
    files: list[UploadFile],
    title_box: str | None,
) -> tuple[list[tuple[str, str]], str | None]:
    for file in files[:1]:
        data = await file.read(state().settings.max_upload_bytes + 1)
        if len(data) > state().settings.max_upload_bytes:
            raise HTTPException(status_code=413, detail="Uploaded image is too large.")
        if not data:
            continue
        try:
            candidates = await asyncio.to_thread(
                _read_inventory_title_candidates,
                data,
                title_box,
            )
            return candidates, None
        except Exception as exc:
            return [], f"Could not read {file.filename or 'image'}: {exc}"
    return [], None


async def _read_inventory_scanner_image(
    files: list[UploadFile],
) -> tuple[bytes, str | None]:
    for file in files[:1]:
        data = await file.read(state().settings.max_upload_bytes + 1)
        if len(data) > state().settings.max_upload_bytes:
            raise HTTPException(status_code=413, detail="Uploaded image is too large.")
        if not data:
            continue
        return data, None
    return b"", None


def _read_inventory_title(
    image_data: bytes,
    title_box: str | None,
) -> tuple[str, str | None, bool]:
    # Tooltip placement can change from one inventory slot to the next. Always
    # check the normal title row first instead of trusting the previous frame's
    # calibration, which may now point at unrelated game UI.
    candidate_boxes = (
        _DEFAULT_INVENTORY_TITLE_BOX,
        title_box,
        *_INVENTORY_TITLE_FALLBACK_BOXES,
    )
    attempted: set[str] = set()
    for candidate_box in candidate_boxes:
        if not candidate_box or candidate_box in attempted:
            continue
        attempted.add(candidate_box)
        fast_text = _read_calibrated_inventory_title(image_data, candidate_box)
        if _inventory_title_candidate_is_plausible(fast_text):
            return fast_text, candidate_box, True
    # Keep live requests bounded to the title-only recognition path. Falling
    # back to full-frame detection takes several seconds and skips items that
    # users hover for the required one-second interval.
    return "", title_box, False


def _read_inventory_title_candidates(
    image_data: bytes,
    title_box: str | None,
) -> list[tuple[str, str]]:
    candidate_boxes = (
        _DEFAULT_INVENTORY_TITLE_BOX,
        title_box,
        *_INVENTORY_TITLE_FALLBACK_BOXES,
    )
    attempted_boxes: set[str] = set()
    seen_text: set[str] = set()
    candidates: list[tuple[str, str]] = []
    for candidate_box in candidate_boxes:
        if not candidate_box or candidate_box in attempted_boxes:
            continue
        attempted_boxes.add(candidate_box)
        text = _read_calibrated_inventory_title(image_data, candidate_box)
        normalized = _normalize_text(text)
        if not _inventory_title_candidate_is_plausible(text) or normalized in seen_text:
            continue
        seen_text.add(normalized)
        candidates.append((text, candidate_box))
    return candidates


def _inventory_title_boxes(title_box: str | None) -> tuple[str, ...]:
    boxes: list[str] = []
    for candidate_box in (
        _DEFAULT_INVENTORY_TITLE_BOX,
        title_box,
    ):
        if candidate_box and candidate_box not in boxes:
            boxes.append(candidate_box)
    return tuple(boxes)


def _inventory_title_candidate_is_plausible(text: str) -> bool:
    normalized = _normalize_text(text)
    compact = normalized.replace(" ", "")
    if len(compact) < 5 or sum(character.isalpha() for character in compact) < 2:
        return False
    scanner_chrome = {
        "star",
        "start",
        "tostar",
        "totostar",
        "focotostat",
    }
    if compact in scanner_chrome:
        return False
    return not _inventory_scanner_line_is_metadata(text)


def _read_inventory_title_bands(image_data: bytes) -> str:
    """Read overlapping horizontal lines without assuming a fixed tooltip position."""
    try:
        from PIL import Image

        image = Image.open(BytesIO(image_data)).convert("RGB")
        width, height = image.size
        if width < 2 or height < 2:
            return ""
        engine = _initialize_rapid_title_ocr()
        band_height = min(height, 32)
        step = 16
        texts: list[str] = []
        seen: set[str] = set()
        for top in range(0, max(1, height - band_height + step), step):
            bottom = min(height, top + band_height)
            if bottom - top < 16:
                continue
            band = image.crop((0, top, width, bottom))
            band = band.resize((width * 2, (bottom - top) * 2))
            output = BytesIO()
            band.save(output, format="PNG")
            result, _ = _run_rapid_title_ocr(output.getvalue(), engine)
            text = " ".join(
                str(item[1]).strip()
                for item in result or []
                if len(item) > 1 and str(item[1]).strip()
            ).strip()
            normalized = _normalize_text(text)
            if (
                not text
                or len(normalized) < 3
                or normalized in seen
                or _inventory_scanner_line_is_metadata(text)
            ):
                continue
            seen.add(normalized)
            texts.append(text)
        return "\n".join(texts)
    except Exception:
        return ""


def _top_inventory_ocr_candidate(result: object) -> Any | None:
    candidates = [
        item for item in result or []
        if len(item) > 1 and str(item[1]).strip()
    ]
    if not candidates:
        return None

    def bounds(item: Any) -> tuple[float, float, float, float]:
        try:
            xs = [float(point[0]) for point in item[0]]
            ys = [float(point[1]) for point in item[0]]
            return min(xs), min(ys), max(xs), max(ys)
        except Exception:
            return float("inf"), float("inf"), float("inf"), float("inf")

    volume_anchors = [
        item for item in candidates
        if re.search(r"\bvol(?:ume)?\s*[:;]?", _normalize_text(str(item[1])), re.IGNORECASE)
    ]
    anchored_titles: list[tuple[float, float, Any]] = []
    for anchor in volume_anchors:
        anchor_left, anchor_top, anchor_right, anchor_bottom = bounds(anchor)
        anchor_height = max(1.0, anchor_bottom - anchor_top)
        for item in candidates:
            if item is anchor or _inventory_scanner_line_is_metadata(str(item[1])):
                continue
            left, top, right, bottom = bounds(item)
            vertical_gap = anchor_top - bottom
            if vertical_gap < -anchor_height or vertical_gap > max(70.0, anchor_height * 5.0):
                continue
            horizontal_overlap = max(0.0, min(right, anchor_right) - max(left, anchor_left))
            left_delta = abs(left - anchor_left)
            if horizontal_overlap <= 0 and left_delta > max(80.0, (anchor_right - anchor_left) * 1.5):
                continue
            anchored_titles.append((
                anchor_top,
                max(0.0, vertical_gap) + (left_delta * 0.12),
                item,
            ))
    if anchored_titles:
        return min(anchored_titles, key=lambda entry: (entry[0], entry[1]))[2]

    candidate = min(candidates, key=lambda item: bounds(item)[1])
    text = str(candidate[1]).strip()
    if _inventory_scanner_line_is_metadata(text):
        return None
    return candidate


def _read_title_above_volume_anchor(
    image_data: bytes,
    result: object,
    selected_candidate: Any,
) -> tuple[str, list[list[float]]] | None:
    anchors: list[tuple[float, Any]] = []
    for item in result or []:
        if len(item) <= 1:
            continue
        if not re.search(r"\bvol(?:ume)?\s*[:;]?", _normalize_text(str(item[1])), re.IGNORECASE):
            continue
        try:
            anchor_top = min(float(point[1]) for point in item[0])
        except Exception:
            continue
        anchors.append((anchor_top, item))
    if not anchors:
        return None

    _, anchor = min(anchors, key=lambda entry: entry[0])
    try:
        anchor_xs = [float(point[0]) for point in anchor[0]]
        anchor_ys = [float(point[1]) for point in anchor[0]]
        candidate_bottom = max(float(point[1]) for point in selected_candidate[0])
        anchor_left = min(anchor_xs)
        anchor_top = min(anchor_ys)
        anchor_height = max(1.0, max(anchor_ys) - anchor_top)
    except Exception:
        return None

    # When detection found a line immediately above Volume, it is already the
    # strongest title read. Recognition-only recovery is for faint/missed titles.
    if -anchor_height <= anchor_top - candidate_bottom <= anchor_height * 1.5:
        text = str(selected_candidate[1]).strip()
        if not text or _inventory_scanner_line_is_metadata(text):
            return None
        return text, [
            [float(point[0]), float(point[1])]
            for point in selected_candidate[0]
        ]

    try:
        from PIL import Image

        image = Image.open(BytesIO(image_data)).convert("RGB")
        width, height = image.size
        left = max(0, round(anchor_left - 6))
        top = max(0, round(anchor_top - (anchor_height * 1.45)))
        right = min(width, round(left + max(300.0, width * 0.24)))
        bottom = min(height, round(anchor_top))
        if right <= left or bottom <= top:
            return None
        crop = image.crop((left, top, right, bottom))
        output = BytesIO()
        crop.save(output, format="PNG")
        recovered, _ = _run_rapid_title_ocr(output.getvalue())
        text = " ".join(
            str(item[1]).strip()
            for item in recovered or []
            if len(item) > 1 and str(item[1]).strip()
        ).strip()
        if not text or _inventory_scanner_line_is_metadata(text):
            return None
        points = [
            [float(left), float(top)],
            [float(right), float(top)],
            [float(right), float(bottom)],
            [float(left), float(bottom)],
        ]
        return text, points
    except Exception:
        return None


def _normalized_ocr_box(image_data: bytes, points: object) -> str | None:
    try:
        from PIL import Image
        image = Image.open(BytesIO(image_data))
        width, height = image.size
        coordinates = list(points)
        ys = [float(point[1]) for point in coordinates]
        padding_y = max(2, (max(ys) - min(ys)) * 0.12)
        left = 0
        top = max(0, min(ys) - padding_y)
        right = width
        bottom = min(height, max(ys) + padding_y)
        values = (left / width, top / height, (right - left) / width, (bottom - top) / height)
        return ",".join(f"{value:.6f}" for value in values)
    except Exception:
        return None


def _read_calibrated_inventory_title(image_data: bytes, title_box: str) -> str:
    try:
        from PIL import Image
        values = [float(value) for value in title_box.split(",")]
        if len(values) != 4 or any(value < 0 or value > 1 for value in values):
            return ""
        image = Image.open(BytesIO(image_data)).convert("RGB")
        width, height = image.size
        left = round(values[0] * width)
        top = round(values[1] * height)
        right = round((values[0] + values[2]) * width)
        bottom = round((values[1] + values[3]) * height)
        if right <= left or bottom <= top:
            return ""
        crop = image.crop((left, top, right, bottom))
        crop = crop.resize((crop.width * 3, crop.height * 3))
        output = BytesIO()
        crop.save(output, format="PNG")
        result, _ = _run_rapid_title_ocr(output.getvalue())
        text = " ".join(
            str(item[1]).strip()
            for item in result or []
            if len(item) > 1 and str(item[1]).strip()
        ).strip()
        # A tight crop can include the beginning of the Volume line. Remove
        # only the characteristic OCR form of its µSCU unit from the end of
        # the title; legitimate model numbers and variant suffixes remain.
        text = re.sub(r"(?i)\s*[o0uµ]*\s*s[c(]u\s*$", "", text).strip()
        return "" if _inventory_scanner_line_is_metadata(text) else text
    except Exception:
        return ""


def _read_image_text(image_data: bytes) -> str:
    rapid_text, rapid_error = _read_image_text_with_rapidocr(image_data)
    if rapid_text.strip() or rapid_error is None:
        return rapid_text

    tesseract_text, tesseract_error = _read_image_text_with_tesseract(image_data)
    if tesseract_text.strip() or tesseract_error is None:
        return tesseract_text
    raise RuntimeError(f"Bundled OCR failed: {rapid_error}. Optional Tesseract fallback failed: {tesseract_error}")


def _read_image_text_with_rapidocr(image_data: bytes) -> tuple[str, str | None]:
    try:
        _initialize_rapid_ocr_pool()
        engine = _RAPID_OCR_POOL.get(timeout=30)
    except Exception as exc:
        return "", str(exc)

    try:
        result, _ = engine(image_data)
        lines = [str(item[1]).strip() for item in result or [] if len(item) > 1 and str(item[1]).strip()]
        return "\n".join(lines), None
    except Exception as exc:
        return "", str(exc)
    finally:
        _RAPID_OCR_POOL.put(engine)


def _initialize_rapid_ocr_pool() -> None:
    """Warm the full OCR engine used to calibrate the title position."""
    global _RAPID_OCR_POOL_READY
    if _RAPID_OCR_POOL_READY:
        return
    from rapidocr_onnxruntime import RapidOCR
    with _RAPID_OCR_LOCK:
        if _RAPID_OCR_POOL_READY:
            return
        while _RAPID_OCR_POOL.qsize() < _RAPID_OCR_POOL_SIZE:
            _RAPID_OCR_POOL.put(RapidOCR())
        _RAPID_OCR_POOL_READY = True


def _initialize_rapid_title_ocr():
    global _RAPID_TITLE_OCR_POOL_READY
    if _RAPID_TITLE_OCR_POOL_READY:
        return None
    from rapidocr_onnxruntime import RapidOCR
    with _RAPID_OCR_LOCK:
        if _RAPID_TITLE_OCR_POOL_READY:
            return None
        while _RAPID_TITLE_OCR_POOL.qsize() < _RAPID_TITLE_OCR_POOL_SIZE:
            _RAPID_TITLE_OCR_POOL.put(RapidOCR(use_text_det=False, use_angle_cls=False))
        _RAPID_TITLE_OCR_POOL_READY = True
    return None


def _warm_rapid_title_ocr() -> None:
    from PIL import Image

    _initialize_rapid_title_ocr()
    image = Image.new("RGB", (320, 48), "black")
    output = BytesIO()
    image.save(output, format="PNG")
    _run_rapid_title_ocr(output.getvalue())


def _run_rapid_title_ocr(image_data: bytes, engine: Any | None = None) -> tuple[Any, Any]:
    """Run recognition with a dedicated pooled engine per concurrent request."""
    if engine is not None:
        return engine(image_data)
    _initialize_rapid_title_ocr()
    title_engine = _RAPID_TITLE_OCR_POOL.get(timeout=30)
    try:
        return title_engine(image_data)
    finally:
        _RAPID_TITLE_OCR_POOL.put(title_engine)


def _rapid_ocr_engine():
    """Retained for compatibility with non-live OCR callers and tests."""
    global _RAPID_OCR
    if _RAPID_OCR is not None:
        return _RAPID_OCR
    from rapidocr_onnxruntime import RapidOCR
    with _RAPID_OCR_LOCK:
        if _RAPID_OCR is None:
            _RAPID_OCR = RapidOCR()
    return _RAPID_OCR


def _read_image_text_with_tesseract(image_data: bytes) -> tuple[str, str | None]:
    try:
        from PIL import Image
        import pytesseract
    except Exception as exc:
        return "", str(exc)

    try:
        image = Image.open(BytesIO(image_data))
        return pytesseract.image_to_string(image), None
    except Exception as exc:
        return "", str(exc)


async def _match_blueprints_from_text(text: str) -> list[dict[str, Any]]:
    matches: dict[str, dict[str, Any]] = {}
    for candidate in _blueprint_text_candidates(text):
        results = await state().sources.lookup_blueprints(query=candidate, limit=5)
        for result in results:
            confidence = _blueprint_match_confidence(candidate, result.name)
            if confidence < 0.58:
                continue
            existing = matches.get(result.name)
            if existing and existing["confidence"] >= confidence:
                continue
            matches[result.name] = {
                "name": result.name,
                "category": result.category,
                "source_name": result.source_name,
                "source_url": result.source_url,
                "component_size": result.component_size,
                "confidence": round(confidence, 2),
                "matched_text": candidate,
            }
    return sorted(matches.values(), key=lambda item: (-item["confidence"], item["name"].lower()))


def _blueprint_text_candidates(text: str) -> list[str]:
    candidates: list[str] = []
    seen: set[str] = set()
    for raw_line in re.split(r"[\r\n]+", text):
        line = _clean_blueprint_ocr_line(raw_line)
        if not line:
            continue
        parts = [line]
        parts.extend(part.strip() for part in re.split(r"\s{2,}|[|•·]", line) if part.strip())
        for part in parts:
            if 3 <= len(part) <= 80 and _normalize_text(part) not in seen:
                seen.add(_normalize_text(part))
                candidates.append(part)
        if len(candidates) >= 120:
            break
    return candidates


def _inventory_items_from_text(
    text: str,
    default_location: str | None = None,
    default_category: str | None = None,
    first_match: bool = False,
) -> list[dict[str, Any]]:
    location = (default_location or "").strip() or "Unknown location"
    category = (default_category or "").strip() or None
    if first_match:
        tooltip_item = _inventory_item_from_tooltip_text(text, location, category)
        return [tooltip_item] if tooltip_item else []

    items: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str | None]] = set()
    for raw_line in re.split(r"[\r\n]+", text):
        parsed = _inventory_item_from_ocr_line(raw_line, location, category)
        if parsed is None:
            continue
        key = (parsed["name"].casefold(), parsed["location"].casefold(), parsed["category"])
        if key in seen:
            continue
        seen.add(key)
        items.append(parsed)
        if first_match:
            break
        if len(items) >= 80:
            break
    return items


async def _enrich_inventory_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    enriched: list[dict[str, Any]] = []
    for item in items:
        enriched_item = item.copy()
        try:
            results = await state().sources.lookup_inventory_items(item["name"], limit=5)
        except Exception:
            results = []
        match = _best_inventory_item_lookup(item["name"], results)
        if match:
            enriched_item["name"] = match.name
            enriched_item["category"] = match.category or enriched_item.get("category")
            enriched_item["item_type"] = (
                match.section
                or _inventory_catalog_item_type(match.name, enriched_item.get("category"))
                or enriched_item.get("item_type")
            )
            enriched_item["item_size"] = match.size or enriched_item.get("item_size")
            enriched_item["source_name"] = match.source_name
            enriched_item["source_url"] = match.source_url
        enriched.append(enriched_item)
    return enriched


async def _match_inventory_scanner_text(
    text: str,
    default_location: str | None,
    default_category: str | None,
    min_score: float,
    exclude_words: str | None,
    scanner_lookups: dict[str, list[tuple[Any, float]]] | None = None,
) -> list[dict[str, Any]]:
    location = (default_location or "").strip() or "Unknown location"
    category = (default_category or "").strip() or None
    matches: dict[str, dict[str, Any]] = {}
    exclude = {_normalize_text(word) for word in re.split(r"[,;\n]+", exclude_words or "") if word.strip()}

    for candidate in _inventory_scanner_text_candidates(text, exclude):
        for result, confidence in _inventory_scanner_accepted_matches(
            scanner_lookups.get(candidate, []) if scanner_lookups is not None else await _inventory_lookup_scored_matches(candidate, 5),
            min_score,
            candidate,
        ):
            existing = matches.get(result.name)
            if existing and existing["confidence"] >= confidence:
                continue
            item = _inventory_item_from_tooltip_text(text, location, category, result.name) or _inventory_item_from_ocr_line(
                result.name,
                location,
                category,
            )
            if item is None:
                continue
            if not _inventory_tooltip_match_agrees_with_result(item, result.name):
                continue
            item.update(
                {
                    "name": result.name,
                    # Keep the website taxonomy selected by the user. Upstream
                    # has additional labels such as Cargo and Usable.
                    "category": category or result.category or item.get("category"),
                    "item_type": (
                        result.section
                        or _inventory_catalog_item_type(result.name, result.category or item.get("category"))
                        or item.get("item_type")
                    ),
                    "item_size": result.size or item.get("item_size"),
                    "source_name": result.source_name,
                    "source_url": result.source_url,
                    "confidence": round(confidence, 2),
                    "matched_text": candidate,
                }
            )
            item["notes"] = _inventory_scanner_notes(item.get("notes"), confidence, candidate)
            matches[result.name] = item

    return sorted(matches.values(), key=lambda item: (-float(item["confidence"]), item["name"].lower()))[:5]


def _inventory_catalog_item_type(name: str, category: str | None) -> str | None:
    normalized = _normalize_text(name)
    word_groups = {
        "Armor": (
            ("Helmet", ("helmet",)),
            ("Torso Armor", ("torso", "chest", "core armor")),
            ("Arm Armor", ("arm armor", "arms")),
            ("Leg Armor", ("leg armor", "legs")),
            ("Backpack", ("backpack",)),
            ("Undersuit", ("undersuit",)),
        ),
        "Clothing": (
            ("Footwear", ("boot", "boots", "shoe", "shoes", "footwear")),
            ("Gloves", ("glove", "gloves")),
            ("Pants", ("pants", "trouser", "trousers")),
            ("Jacket", ("jacket", "coat")),
            ("Shirt", ("shirt", "sweater")),
            ("Hat", ("hat", "cap", "head gear", "headgear")),
        ),
        "Weapons": (
            ("Attachments", ("scope", "optic", "suppressor", "compensator", "attachment")),
            ("Sidearm", ("pistol", "sidearm")),
            ("Melee", ("knife", "melee")),
            ("Primary", ("rifle", "shotgun", "smg", "lmg", "sniper", "launcher", "railgun")),
        ),
        "Utility": (
            ("Medical", ("medical", "medpen", "med pen", "paramed")),
            ("Multitool", ("multi tool", "multitool")),
            ("Mining", ("mining",)),
            ("Salvage", ("salvage",)),
            ("Container", ("container",)),
            ("Tool", ("tool", "tractor beam")),
        ),
        "Ammunition": (
            ("Magazine", ("magazine",)),
            ("Battery", ("battery",)),
            ("Ammunition", ("ammunition", "ammo")),
        ),
        "Components": (
            ("Ordnance", ("ordnance", "bomb", "missile", "torpedo", "rocket pod")),
            ("Coolers", ("cooler",)),
            ("EMP", ("emp",)),
            ("Mining", ("mining",)),
            ("Missile Racks", ("missile rack",)),
            ("Power Plants", ("power plant", "powerplant")),
            ("Quantum Drives", ("quantum drive",)),
            ("Shields", ("shield generator", "shield")),
            ("Turrets", ("turret",)),
            ("Weapons", ("vehicle weapon", "ship weapon", "cannon", "repeater")),
            ("CM Launchers", ("cm launcher", "countermeasure launcher")),
            ("Liveries", ("livery", "paint")),
            ("Jump Modules", ("jump module", "jump drive")),
            ("Radar", ("radar",)),
        ),
        "Sustenance": (
            ("Drink", ("drink", "water", "soda")),
            ("Food", ("food",)),
        ),
        "Other": (
            ("Paint", ("paint",)),
            ("Flair", ("flair",)),
            ("Collectible", ("collectible",)),
            ("Container", ("container",)),
        ),
    }
    for item_type, terms in word_groups.get(category or "", ()):
        if any(term in normalized for term in terms):
            return item_type
    return None


def _inventory_manual_catalog_category(result: Any) -> str | None:
    source_category = _normalize_text(getattr(result, "category", "") or "")
    section = _normalize_text(getattr(result, "section", "") or "")
    combined = f"{source_category} {section}"
    direct_markers = (
        ("Armor", ("armor", "armour", "undersuit", "backpack")),
        ("Clothing", ("clothing", "clothes")),
        ("Ammunition", ("ammunition", "magazine", "ammo")),
        ("Sustenance", ("food", "drink", "sustenance")),
        ("Utility", ("usable", "utility", "multitool", "tractor beam", "medical")),
        ("Components", ("vehicle component", "cooler", "power plant", "quantum drive", "shield")),
        ("Weapons", ("personal weapon", "weaponpersonal")),
    )
    for category, markers in direct_markers:
        if any(marker in combined for marker in markers):
            return category
    name = str(getattr(result, "name", "") or "")
    for category in ("Armor", "Clothing", "Weapons", "Utility", "Ammunition", "Components", "Sustenance", "Other"):
        if _inventory_catalog_item_type(name, category):
            return category
    return None


def _inventory_scanner_accepted_matches(
    scored_matches: list[tuple[Any, float]],
    min_score: float,
    candidate: str | None = None,
) -> list[tuple[Any, float]]:
    # A long tooltip title with a clearly separated catalog winner remains safe
    # slightly below the live threshold. This recovers screen-share OCR damage
    # without weakening short or ambiguous names.
    ranked = sorted(scored_matches, key=lambda item: (-item[1], -len(item[0].name), item[0].name.lower()))
    adaptive_floor = min_score
    if candidate and ranked:
        normalized_candidate = _normalize_text(_normalize_inventory_tooltip_name(candidate))
        runner_up = ranked[1][1] if len(ranked) > 1 else 0
        if len(normalized_candidate) >= 14 and ranked[0][1] >= 0.82 and ranked[0][1] - runner_up >= 0.05:
            adaptive_floor = 0.82
    accepted = sorted(
        [(result, confidence) for result, confidence in scored_matches if confidence >= adaptive_floor],
        key=lambda item: (-item[1], -len(item[0].name), item[0].name.lower()),
    )
    if not accepted:
        return []
    if candidate:
        candidate_words = _normalize_text(_normalize_inventory_tooltip_name(candidate)).split()
        result_words = _normalize_text(accepted[0][0].name).split()
        if len(candidate_words) >= 3 and len(result_words) >= 3:
            family_similarity = difflib.SequenceMatcher(
                None, candidate_words[0], result_words[0]
            ).ratio()
            # CamelCase catalog families such as RediMake may be split by OCR
            # normalization. Compare the joined leading token as well.
            family_similarity = max(
                family_similarity,
                difflib.SequenceMatcher(
                    None, "".join(candidate_words[:2]), result_words[0]
                ).ratio(),
            )
            if family_similarity < 0.70:
                return []
    if len(accepted) > 1 and accepted[0][1] - accepted[1][1] < 0.04:
        if not candidate or not _inventory_top_match_has_distinctive_candidate_word(
            candidate, accepted[0][0].name, accepted[1][0].name
        ):
            return []
    return [accepted[0]]


def _inventory_top_match_has_distinctive_candidate_word(
    candidate: str, top_name: str, runner_up_name: str
) -> bool:
    words = lambda value: set(re.findall(r"[a-z0-9]+", _normalize_text(value)))
    candidate_words = words(_normalize_inventory_tooltip_name(candidate))
    top_words = words(top_name)
    runner_up_words = words(runner_up_name)
    distinctive = {
        word for word in top_words - runner_up_words if len(word) >= 3 and not word.isdigit()
    }
    return bool(candidate_words & distinctive)


def _inventory_tooltip_match_agrees_with_result(item: dict[str, Any], result_name: str) -> bool:
    item_name = _normalize_text(str(item.get("name") or ""))
    result = _normalize_text(result_name)
    attachment_terms = {"compensator", "suppressor"}
    item_terms = {term for term in attachment_terms if term in item_name}
    result_terms = {term for term in attachment_terms if term in result}
    if item_terms and result_terms and item_terms.isdisjoint(result_terms):
        return False
    return True


async def _inventory_scanner_diagnostics(
    text: str,
    min_score: float,
    exclude_words: str | None,
    scanner_lookups: dict[str, list[tuple[Any, float]]] | None = None,
) -> dict[str, Any]:
    exclude = {_normalize_text(word) for word in re.split(r"[,;\n]+", exclude_words or "") if word.strip()}
    raw_lines = [_clean_inventory_ocr_line(line) for line in re.split(r"[\r\n]+", text)]
    raw_lines = [line for line in raw_lines if line]
    candidate_values = _inventory_scanner_text_candidates(text, exclude)
    candidates: list[dict[str, Any]] = []

    for candidate in candidate_values[:30]:
        try:
            results = scanner_lookups.get(candidate, []) if scanner_lookups is not None else await _inventory_lookup_scored_matches(candidate, 5)
        except Exception as exc:
            candidates.append(
                {
                    "text": candidate,
                    "status": "lookup_error",
                    "reason": str(exc),
                    "matches": [],
                }
            )
            continue

        scored_matches = []
        for result, score in results:
            scored_matches.append(
                {
                    "name": result.name,
                    "category": result.category,
                    "item_type": result.section,
                    "size": result.size,
                    "source_name": result.source_name,
                    "source_url": result.source_url,
                    "score": round(score, 2),
                    "accepted": score >= min_score,
                }
            )
        best_score = max((float(match["score"]) for match in scored_matches), default=0.0)
        candidates.append(
            {
                "text": candidate,
                "status": "accepted" if best_score >= min_score else "rejected",
                "reason": "catalog score passed" if best_score >= min_score else f"best score below {min_score:g}",
                "matches": sorted(scored_matches, key=lambda match: -float(match["score"])),
            }
        )

    candidate_norms = {_normalize_text(candidate) for candidate in candidate_values}
    rejected_lines = [
        {
            "text": line,
            "reason": "metadata/noise" if _inventory_scanner_line_is_metadata(line) else "not selected as candidate",
        }
        for line in raw_lines[:80]
        if _normalize_text(line) not in candidate_norms
    ]
    return {
        "min_score": min_score,
        "candidate_count": len(candidate_values),
        "candidates": candidates,
        "rejected_lines": rejected_lines[:40],
    }


async def _inventory_scanner_lookups(
    text: str,
    exclude_words: str | None,
    candidate_limit: int | None = None,
    category: str | None = None,
    item_type: str | None = None,
) -> dict[str, list[tuple[Any, float]]]:
    """Resolve each OCR candidate once for both matching and diagnostics.

    Catalog lookups are network-bound. A small concurrency limit keeps a scan
    inside reverse-proxy timeouts without flooding the upstream data source.
    """
    exclude = {_normalize_text(word) for word in re.split(r"[,;\n]+", exclude_words or "") if word.strip()}
    candidates = _inventory_scanner_text_candidates(text, exclude)
    if candidate_limit is not None:
        candidates = candidates[:max(1, candidate_limit)]
    semaphore = asyncio.Semaphore(4)

    async def lookup(candidate: str) -> tuple[str, list[tuple[Any, float]]]:
        async with semaphore:
            matches = (
                await _inventory_lookup_scored_matches(candidate, 5, category, quick=True)
                if category
                else await _inventory_lookup_scored_matches(candidate, 5, quick=True)
            )
            selected_type = _normalize_text(item_type or "")
            if selected_type:
                typed_matches = [
                    (result, score)
                    for result, score in matches
                    if selected_type in {
                        _normalize_text(getattr(result, "section", "") or ""),
                        _normalize_text(
                            _inventory_catalog_item_type(
                                getattr(result, "name", ""),
                                getattr(result, "category", None) or category,
                            )
                            or ""
                        ),
                    }
                ]
                if typed_matches:
                    matches = typed_matches
            return candidate, matches

    return dict(await asyncio.gather(*(lookup(candidate) for candidate in candidates)))


async def _inventory_lookup_scored_matches(
    candidate: str,
    limit: int = 5,
    category: str | None = None,
    quick: bool = False,
) -> list[tuple[Any, float]]:
    seen: set[str] = set()
    scored: list[tuple[Any, float]] = []

    async def lookup(query: str) -> list[Any]:
        try:
            categorized = (
                await state().sources.lookup_inventory_items(
                    query, limit=limit, category=category
                )
                if category
                else []
            )
            if any(
                max(
                    _inventory_match_confidence(candidate, name)
                    for name in (
                        result.name,
                        *getattr(result, "catalog_aliases", ()),
                    )
                )
                >= 0.88
                for result in categorized
            ):
                return categorized
            # Source categories do not map one-to-one to the website taxonomy
            # (for example Cargo -> Commodities and Usable -> Utility). Search
            # the complete catalog too, then rely on confidence and ambiguity
            # checks instead of silently hiding valid items.
            complete = await state().sources.lookup_inventory_items(query, limit=limit)
            combined: list[Any] = []
            combined_names: set[str] = set()
            for result in [*categorized, *complete]:
                key = _normalize_text(getattr(result, "name", ""))
                if not key or key in combined_names:
                    continue
                combined_names.add(key)
                combined.append(result)
            return combined[: limit * 2]
        except Exception:
            return []

    queries = _inventory_lookup_queries(candidate)
    if quick:
        queries = queries[:2]
    result_groups = await asyncio.gather(*(lookup(query) for query in queries))
    # Prefer supplements so their OCR aliases are retained when an incomplete
    # upstream result has the same canonical name.
    supplements = _inventory_scanner_catalog_supplements(candidate)
    result_groups.insert(0, supplements)
    existing_results = [result for group in result_groups for result in group]
    existing_is_confident = any(
        max(
            _inventory_match_confidence(candidate, name)
            for name in (result.name, *getattr(result, "catalog_aliases", ()))
        )
        >= 0.88
        for result in existing_results
    )
    # Data.p4k closes catalog/version gaps. Keep the richer Wiki result when
    # it already matches confidently, and pay the P4K fuzzy-index cost only
    # when the external catalog cannot resolve the OCR candidate.
    if not existing_is_confident:
        result_groups.append(P4K_INVENTORY_CATALOG.lookup(candidate, limit * 2))
    for results in result_groups:
        for result in results:
            key = _normalize_text(getattr(result, "name", ""))
            if not key or key in seen:
                continue
            seen.add(key)
            names = (result.name, *getattr(result, "catalog_aliases", ()))
            scored.append(
                (
                    result,
                    max(_inventory_match_confidence(candidate, name) for name in names),
                )
            )
    return sorted(scored, key=lambda item: (-item[1], item[0].name.lower()))[:limit]


def _inventory_scanner_catalog_supplements(candidate: str) -> list[ItemLocatorResult]:
    """Supply confirmed variants missing from the upstream searchable item catalog."""
    normalized = _normalize_text(_normalize_inventory_tooltip_name(candidate))
    variants = (
        (("bloodline",), 'Pyro RYT "Bloodline" Multi-Tool', "Multitool", "Utility", "Greycat Industrial"),
        (("hurston",), 'Pyro RYT "Hurston" Multi-Tool', "Multitool", "Utility", "Greycat Industrial"),
        (("microtech",), 'Pyro RYT "microTech" Multi-Tool', "Multitool", "Utility", "Greycat Industrial"),
        (("xdl", "mark", "monocular", "rangefinder"), 'XDL "Mark I" Monocular Rangefinder', "Gadgets", "Utility", "Behring Applied Technology"),
        (("maxlift", "aa", "support", "tractor", "beam"), "MaxLift AA Support Tractor Beam", "Tractor Beams", "Utility", "Greycat Industrial"),
        (("maxlift", "aa", "transport", "tractor", "beam"), "MaxLift AA Transport Tractor Beam", "Tractor Beams", "Utility", "Greycat Industrial"),
        (("tumbril", "cargo", "plushie"), "Tumbril Cargo Plushie", "Flair", "Other", "Tumbril Land Systems"),
        (("redimake", "item", "fabricator", "aa", "support"), "RediMake Item Fabricator AA Support", "Crafter", "Other", "RediMake"),
    )
    results: list[ItemLocatorResult] = []
    compact = normalized.replace(" ", "")
    if compact in {"tractorbeam", "xtractorbeam", "extractorbeam"}:
        results.append(
            ItemLocatorResult(
                id=-9_999,
                name="TruHold Tractor Beam Attachment",
                section="Attachments",
                category="Utility",
                company_name="Greycat Industrial",
                size=None,
                wiki_url="https://starcitizen.tools/TruHold_Tractor_Beam_Attachment",
                source_url="https://starcitizen.tools/TruHold_Tractor_Beam_Attachment",
                source_name="Star Citizen Wiki",
                purchases=[],
                catalog_aliases=("Tractorbeam", "Tractor Beam"),
            )
        )
    for markers, name, section, category, company in variants:
        if not all(marker in compact for marker in markers):
            continue
        slug = re.sub(r"[^a-z0-9]+", "_", name.casefold()).strip("_")
        wiki_title = name.replace(" ", "_").replace('"', "%22")
        results.append(
            ItemLocatorResult(
                id=-(10_000 + len(results)),
                name=name,
                section=section,
                category=category,
                company_name=company,
                size=None,
                wiki_url=f"https://starcitizen.tools/{wiki_title}",
                source_url=f"https://starcitizen.tools/{wiki_title}",
                source_name="Star Citizen Wiki",
                purchases=[],
                catalog_aliases=(f"pyro_ryt_{slug}_multitool",),
            )
        )
    return results


def _inventory_lookup_queries(candidate: str) -> list[str]:
    normalized = _normalize_inventory_tooltip_name(candidate)
    words = [word for word in re.split(r"\s+", normalized.strip()) if word]
    queries = [normalized]
    if len(words) >= 2:
        queries.append(" ".join(words[:2]))
        if len(words[0]) >= 5 or re.search(r"\d", words[0]):
            queries.append(words[0])
    if len(words) >= 3:
        queries.append(" ".join(words[:3]))
        queries.append(" ".join(words[-2:]))
    return _unique_preserve_order(query for query in queries if len(query) >= 3)


def _inventory_scanner_notes(notes: object, confidence: float, candidate: str) -> str:
    base = str(notes or "Imported from hover scanner")
    suffix = f"Match: {round(confidence * 100)}% from '{candidate}'"
    if suffix in base:
        return base
    return f"{base} ({suffix})"


def _inventory_scanner_text_candidates(text: str, exclude_words: set[str] | None = None) -> list[str]:
    exclude_words = exclude_words or set()
    candidates: list[str] = []
    seen: set[str] = set()
    raw_lines = [line.strip() for line in re.split(r"[\r\n]+", text) if line.strip()]
    # Some game builds expose an internal localization key instead of a
    # display title. Preserve its separators for class-name alias matching;
    # generic OCR cleanup would otherwise lower an otherwise exact match.
    for raw_line in raw_lines:
        if re.search(r"(?i)item[\W_]*name", raw_line):
            _add_inventory_candidate(candidates, seen, raw_line, exclude_words)
    lines = [_clean_inventory_ocr_line(line) for line in raw_lines]
    lines = [line for line in lines if line]
    blocks = _inventory_tooltip_blocks(lines)

    for block in blocks:
        tooltip_name = _inventory_tooltip_name(block)
        if tooltip_name:
            _add_inventory_candidate(candidates, seen, tooltip_name, exclude_words)

    relevant_lines = [line for block in blocks for line in block]
    for line in relevant_lines:
        if _inventory_scanner_line_is_metadata(line):
            continue
        normalized_line = _normalize_inventory_tooltip_name(line)
        _add_inventory_candidate(candidates, seen, normalized_line, exclude_words)
        for part in re.split(r"\s{2,}|[|•·]", normalized_line):
            _add_inventory_candidate(candidates, seen, part, exclude_words)

    joined = " ".join(relevant_lines)
    for pattern in (
        r"\b[A-Z]{2,4}[- ]\d+\s+[A-Z][A-Za-z0-9'\"() -]{2,40}",
        r"\b[A-Z][A-Za-z]+(?:\s+[A-Z][A-Za-z0-9'\"() -]+){1,5}\b",
    ):
        for match in re.finditer(pattern, joined):
            _add_inventory_candidate(candidates, seen, _normalize_inventory_tooltip_name(match.group(0)), exclude_words)
            if len(candidates) >= 30:
                return candidates
    return candidates[:30]


def _add_inventory_candidate(
    candidates: list[str],
    seen: set[str],
    value: str,
    exclude_words: set[str],
) -> None:
    candidate = " ".join(str(value or "").split()).strip(" -:,.")
    normalized = _normalize_text(candidate)
    if not normalized or normalized in seen:
        return
    if len(candidate) < 3 or len(candidate) > 80:
        return
    if any(word and word in normalized for word in exclude_words):
        return
    if _inventory_scanner_line_is_metadata(candidate):
        return
    if not any(char.isalpha() for char in candidate):
        return
    seen.add(normalized)
    candidates.append(candidate)


def _inventory_scanner_line_is_metadata(line: str) -> bool:
    normalized = _normalize_text(line)
    if not normalized:
        return True
    compact = normalized.replace(" ", "")
    blocked_exact = {
        "inventory",
        "personal all",
        "personal backpack",
        "local local",
        "looting view",
        "move all",
        "clear filters",
        "empty",
    }
    blocked_compact = {
        "inventory",
        "personalall",
        "personalbackpack",
        "locallocal",
        "lootingview",
        "moveall",
        "clearfilters",
        "empty",
    }
    if normalized in blocked_exact or compact in blocked_compact:
        return True
    if compact.startswith("move") and len(compact) <= 8:
        return True
    blocked_prefixes = (
        "volume",
        "capacity",
        "manufacturer",
        "type",
        "item type",
        "class",
        "magazine size",
        "rate of fire",
        "fire rate",
        "effective range",
        "attachments",
        "attachment point",
        "underbarrel",
        "barrel",
        "optics",
        "magnification",
        "zoom",
        "aim time",
        "impact force",
        "recoil",
        "damage",
        "size",
        "quality",
    )
    if any(
        re.match(rf"^{re.escape(prefix)}(?:\b|\s|:)", normalized)
        for prefix in blocked_prefixes
    ):
        return True
    if re.fullmatch(r"\d+(?:\.\d+)?\s*(?:scu|uscu|q|rpm|m)", normalized):
        return True
    words = normalized.split()
    if len(words) > 10:
        return True
    if len(words) >= 9 and sum(1 for word in words if word in {"the", "and", "to", "of", "for", "that", "with"}) >= 3:
        return True
    return False


def _inventory_match_confidence(candidate: str, item_name: str) -> float:
    if _inventory_scanner_line_is_metadata(candidate):
        return 0
    candidate_norm = _normalize_text(_normalize_inventory_tooltip_name(candidate))
    item_norm = _normalize_text(_strip_inventory_item_prefix(item_name))
    if not candidate_norm or not item_norm:
        return 0
    if candidate_norm == item_norm:
        return 1
    suffix_score = _inventory_distinctive_suffix_score(candidate_norm, item_norm)
    if item_norm in candidate_norm:
        extra_words = set(candidate_norm.split()) - set(item_norm.split())
        ceiling = 0.88 if extra_words else 0.96
        containment_score = min(
            ceiling,
            len(item_norm) / max(len(candidate_norm), 1) + 0.25,
        )
        if suffix_score >= 0.82:
            return max(containment_score, min(0.99, 0.9 + ((suffix_score - 0.82) * 0.5)))
        return containment_score
    if candidate_norm in item_norm:
        return min(0.9, len(candidate_norm) / max(len(item_norm), 1) + 0.2)
    candidate_words = set(candidate_norm.split())
    item_words = set(item_norm.split())
    overlap = len(candidate_words & item_words)
    word_score = (
        (2 * overlap) / max(len(candidate_words) + len(item_words), 1)
        if overlap else 0
    )
    typo_score = difflib.SequenceMatcher(None, candidate_norm, item_norm).ratio()
    compact_typo_score = difflib.SequenceMatcher(
        None,
        candidate_norm.replace(" ", ""),
        item_norm.replace(" ", ""),
    ).ratio()
    score = max(word_score, typo_score * 0.95, compact_typo_score * 0.99)
    if suffix_score >= 0.82:
        score = max(score, min(0.99, 0.9 + ((suffix_score - 0.82) * 0.5)))
    candidate_family = _inventory_name_family(candidate_norm)
    item_family = _inventory_name_family(item_norm)
    if candidate_family and item_family and candidate_family != item_family:
        family_similarity = difflib.SequenceMatcher(None, candidate_family, item_family).ratio()
        if family_similarity < 0.72 and compact_typo_score < 0.85 and suffix_score < 0.82:
            score = min(score, 0.65)
    candidate_numbers = set(re.findall(r"\d+", candidate_norm))
    item_numbers = set(re.findall(r"\d+", item_norm))
    if candidate_numbers and item_numbers and not (candidate_numbers & item_numbers):
        score = min(score, 0.65)
    # Stabilize threshold comparisons such as an intended 0.88 score that
    # binary floating point otherwise represents as 0.8799999999999999.
    return round(score, 6)


def _inventory_distinctive_suffix_score(candidate_norm: str, item_norm: str) -> float:
    item_words = item_norm.split()
    if not item_words:
        return 0
    # Multi-word catalog names frequently share generic type suffixes such as
    # "Rifle" or "Multi-Tool". Those suffixes cannot identify a variant.
    if len(item_words) > 1:
        return 0
    suffix = item_words[-1]
    if len(suffix) < 5 or not suffix.isalpha():
        return 0
    candidate_words = candidate_norm.split()
    candidate_tail = candidate_words[-1] if candidate_words else candidate_norm
    compact_candidate = candidate_norm.replace(" ", "")
    trailing_scores = [
        difflib.SequenceMatcher(
            None,
            compact_candidate[-length:],
            suffix,
        ).ratio()
        for length in range(max(3, len(suffix) - 2), len(suffix) + 3)
        if len(compact_candidate) >= length
    ]
    return max(
        [difflib.SequenceMatcher(None, candidate_tail, suffix).ratio(), *trailing_scores]
    )


def _inventory_name_family(normalized_name: str) -> str | None:
    for word in normalized_name.split():
        if any(char.isalpha() for char in word) and len(word) >= 3:
            return word
    return None


def _best_inventory_item_lookup(name: str, results: list[Any]) -> Any | None:
    if not results:
        return None
    normalized = _normalize_text(name)
    for result in results:
        if _normalize_text(getattr(result, "name", "")) == normalized:
            return result
    for result in results:
        result_name = _normalize_text(getattr(result, "name", ""))
        if normalized in result_name or result_name in normalized:
            return result
    return results[0]


def _inventory_item_from_tooltip_text(
    text: str,
    default_location: str,
    default_category: str | None,
    matched_name: str | None = None,
) -> dict[str, Any] | None:
    lines = [_clean_inventory_ocr_line(line) for line in re.split(r"[\r\n]+", text)]
    lines = [line for line in lines if line]
    blocks = _inventory_tooltip_blocks(lines)
    lines = _inventory_tooltip_block_for_match(blocks, matched_name) if blocks else []
    if not lines:
        return None
    name = _inventory_tooltip_name(lines)
    if not name:
        return None
    category, item_type = _inventory_tooltip_category(lines, default_category)
    item_size = _inventory_tooltip_size(lines)
    if re.search(r"(?:^|\s)x?\s*\d+(?:\.\d+)?\s*$", lines[0], flags=re.IGNORECASE):
        stack_item = _inventory_item_from_ocr_line(lines[0], default_location, default_category)
        if stack_item:
            stack_item["notes"] = "Imported from hover scanner"
            stack_item["category"] = category or stack_item["category"]
            stack_item["item_type"] = item_type or stack_item["item_type"]
            stack_item["item_size"] = item_size
            return stack_item
    quality = _inventory_tooltip_quality(lines, name)
    volume_scu = _inventory_tooltip_volume_scu(lines)
    uses_scu = _inventory_tooltip_uses_scu(category, item_type, lines)
    if not uses_scu:
        volume_scu = None
        if not _inventory_tooltip_has_explicit_quality(lines):
            quality = None
    if quality is None and volume_scu is None:
        stack_item = _inventory_item_from_ocr_line(name, default_location, default_category)
        if stack_item:
            stack_item["notes"] = "Imported from hover scanner"
            stack_item["category"] = category or stack_item["category"]
            stack_item["item_type"] = item_type or stack_item["item_type"]
            stack_item["item_size"] = item_size
            return stack_item
    details = []
    if quality is not None:
        details.append(f"Quality: {quality:g}")
    if item_size:
        details.append(f"Size: {item_size}")
    if volume_scu is not None:
        details.append(f"Volume: {volume_scu:g} SCU")
    return {
        "name": name,
        "category": category,
        "item_type": item_type,
        "item_size": item_size,
        "location": default_location,
        "quantity": volume_scu if uses_scu and volume_scu is not None else 1.0,
        "quality": quality,
        "volume_scu": volume_scu,
        "notes": "Imported from hover scanner" + (f" ({', '.join(details)})" if details else ""),
    }


def _inventory_tooltip_block_for_match(blocks: list[list[str]], matched_name: str | None) -> list[str]:
    if not blocks:
        return []
    if not matched_name:
        return blocks[0]
    return max(
        blocks,
        key=lambda block: _inventory_match_confidence(_inventory_tooltip_name(block) or "", matched_name),
    )


def _inventory_tooltip_blocks(lines: list[str]) -> list[list[str]]:
    blocks: list[list[str]] = []
    current: list[str] = []
    for line in lines:
        normalized = _normalize_text(line)
        if not normalized:
            continue
        if current and _inventory_line_starts_tooltip(line) and any(_inventory_line_is_tooltip_stat(item) for item in current):
            blocks.append(_inventory_tooltip_relevant_lines(current))
            current = [line]
            continue
        current.append(line)
    if current:
        blocks.append(_inventory_tooltip_relevant_lines(current))
    stat_blocks = [
        block
        for index, block in enumerate(blocks)
        if any(_inventory_line_is_tooltip_stat(line) for line in block)
        and (index == 0 or sum(1 for line in block if _inventory_line_is_tooltip_stat(line)) >= 2)
    ]
    return stat_blocks or blocks


def _inventory_line_starts_tooltip(line: str) -> bool:
    normalized = _normalize_text(line)
    if not normalized or _inventory_scanner_line_is_metadata(line):
        return False
    if _inventory_line_is_tooltip_stat(line):
        return False
    if re.search(r"\b\d{1,4}\b.*\b(?:scu|uscu|5cu)\b", normalized):
        return False
    words = normalized.split()
    if len(words) > 7:
        return False
    if sum(1 for word in words if word in {"the", "and", "to", "of", "for", "that", "with", "where", "you"}) >= 2:
        return False
    return any(char.isalpha() for char in line)


def _inventory_line_is_tooltip_stat(line: str) -> bool:
    normalized = _normalize_text(line)
    return normalized.startswith(
        (
            "volume",
            "manufacturer",
            "type",
            "item type",
            "attachment point",
            "attachmentpoint",
            "magnification",
            "zoom",
            "aim time",
            "parallax",
            "size",
            "class",
            "magazine size",
            "rate of fire",
            "fire rate",
            "effective range",
            "attachments",
            "capacity",
            "quality",
            "visual recoil",
        )
    )


def _inventory_tooltip_relevant_lines(lines: list[str]) -> list[str]:
    relevant: list[str] = []
    saw_stat = False
    after_description = False
    for line in lines:
        normalized = _normalize_text(line)
        if not normalized:
            continue
        if after_description:
            if normalized.startswith(("capacity", "quality")):
                relevant.append(line)
                continue
            if re.search(r"\b\d{1,4}\b.*\b(?:scu|uscu|5cu)\b", normalized) and not normalized.startswith(
                ("volume", "manufacturer", "type", "item type", "attachment", "magnification", "zoom", "aim", "size")
            ):
                relevant.append(line)
            continue
        if _inventory_line_is_tooltip_stat(line):
            saw_stat = True
            relevant.append(line)
            continue
        words = normalized.split()
        looks_like_description = len(words) >= 7 and sum(
            1 for word in words if word in {"the", "and", "to", "of", "for", "that", "with", "where", "you"}
        ) >= 2
        if saw_stat and looks_like_description:
            after_description = True
            continue
        relevant.append(line)
    return relevant


def _inventory_tooltip_name(lines: list[str]) -> str | None:
    blocked_prefixes = (
        "volume",
        "capacity",
        "storage",
        "quality",
        "1:",
        "2:",
        "3:",
        "4:",
        "5:",
        "featuring",
        "the ",
        "this ",
        "when ",
        "it ",
    )
    blocked_exact = {"empty", "personal", "backpack", "local", "looting view", "move all", "tart", "start"}
    for line in lines:
        lowered = line.casefold()
        if lowered in blocked_exact or lowered.startswith(blocked_prefixes):
            continue
        if _inventory_is_noisy_header(line):
            continue
        if re.search(r"\d+\s*(?:/|scu|µscu|uscu|q\b)", lowered):
            continue
        if 3 <= len(line) <= 80 and any(char.isalpha() for char in line):
            return _normalize_inventory_tooltip_name(line)
    return None


def _inventory_tooltip_quality(lines: list[str], name: str) -> float | None:
    for line in lines:
        if "quality" in line.casefold():
            match = re.search(r"(\d+(?:\.\d+)?)", line)
            if match:
                return float(match.group(1))
    label = name.split()[-1] if name else ""
    for index, line in enumerate(lines):
        if label and label.casefold() not in line.casefold():
            if re.fullmatch(r"\d{1,4}", line):
                nearby = " ".join(lines[index + 1 : index + 3]).casefold()
                if label.casefold() in nearby or "scu" in nearby:
                    return float(line)
            continue
        line_without_scu = re.sub(r"\d+(?:\.\d+)?\s*(?:scu|µscu|uscu|5cu)", " ", line, flags=re.IGNORECASE)
        numbers = [
            float(value)
            for value in re.findall(r"(?<![\d.])(\d{1,4})(?![\d.])", line_without_scu)
        ]
        if numbers:
            return numbers[-1]
    return None


def _inventory_tooltip_category(lines: list[str], default_category: str | None) -> tuple[str | None, str | None]:
    joined = " ".join(lines).casefold()
    if re.search(r"\bitem\s*type\s*[: ]\s*(?:lmg|rifle|shotgun|sniper|smg|launcher|railgun)\b", joined):
        return "Personal Weapons", "Primary"
    if re.search(r"\bitem\s*type\s*[: ]\s*(?:pistol|sidearm)\b", joined):
        return "Personal Weapons", "Sidearm"
    if "personal weapon" in joined or re.search(r"\b(pistol|rifle|shotgun|sniper|smg|lmg)\b", joined):
        return "Personal Weapons", "Primary"
    if (
        "attachmentpoint" in joined
        or "attachment point" in joined
        or "magnification" in joined
        or "optic" in joined
        or "holographic" in joined
        or "telescopic" in joined
    ):
        return "Personal Weapons", "Attachments"
    return default_category, None


def _inventory_tooltip_uses_scu(category: str | None, item_type: str | None, lines: list[str]) -> bool:
    joined = " ".join(lines).casefold()
    if category == "Personal Weapons" or item_type in {"Attachments", "Weapons"}:
        return False
    return "scu" in joined or "commodity" in joined or "material" in joined


def _inventory_tooltip_has_explicit_quality(lines: list[str]) -> bool:
    return any("quality" in line.casefold() for line in lines)


def _inventory_tooltip_volume_scu(lines: list[str]) -> float | None:
    joined = " ".join(lines)
    scu_matches = re.findall(r"(\d+(?:\.\d+)?)\s*(?:SCU|5CU)", joined, flags=re.IGNORECASE)
    if scu_matches:
        return float(scu_matches[-1])
    micro_match = re.search(r"Volume\s*[: ]\s*(\d+(?:\.\d+)?)\s*(?:µSCU|uSCU|USCU|pSCU)", joined, flags=re.IGNORECASE)
    if micro_match:
        return float(micro_match.group(1)) / 1_000_000
    return None


def _inventory_tooltip_size(lines: list[str]) -> str | None:
    for line in lines:
        match = re.search(r"^Size\s*[: ]\s*(S?\d+)\b", line, flags=re.IGNORECASE)
        if match:
            value = match.group(1).upper()
            return value if value.startswith("S") else f"Size {value}"
    return None


def _clean_inventory_ocr_line(value: str) -> str:
    value = value.replace("μ", "µ").replace("Âµ", "µ")
    value = re.sub(r"[^A-Za-z0-9'’+./():µ -]", " ", value)
    return " ".join(value.split()).strip(" -.")


def _normalize_inventory_tooltip_name(value: str) -> str:
    value = _strip_inventory_item_prefix(value)
    value = re.sub(r"([a-z])([A-Z])", r"\1 \2", value)
    value = value.replace("KopionHorn", "Kopion Horn")
    value = re.sub(r'(\w)"', r'\1 "', value)
    value = re.sub(r'"(\()', r'" \1', value)
    value = re.sub(r"\)(\w)", r") \1", value)
    value = re.sub(r"(\d+)x([A-Z])", r"\1x \2", value)
    value = " ".join(value.split())
    replacements = {
        r"\bmed\s+pen\b": "MedPen",
        r"\bchemozaly\b": "Hemozal",
        r"\bchemozan\b": "Hemozal",
        r"\bhemozan\b": "Hemozal",
        r"\bbloodino\b": "Bloodline",
        r"\bbloodlino\b": "Bloodline",
        r"\bpiconlla\b": "Piconalia",
        r"\barelight\b": "Arclight",
        r"\brangetinder\b": "Rangefinder",
        r"\bx\s*dl[- ]*mark\b": "XDL Mark",
        r"\bx\s*l[- ]*mark\b": "XDL Mark",
        r"\bpyro\s*ryt\b": "Pyro RYT",
        r"\bryt\s*(bloodline|hurston|micro\s*tech)\b": r"RYT \1",
        r"\bmulti[- ]?tooi\b": "Multi-Tool",
        r"\bmult[- ]?tooi\b": "Multi-Tool",
        r"\bmuti[- ]?tool\b": "Multi-Tool",
        r"\bmult[- ]?tol\b": "Multi-Tool",
        r"\bmut[- ]?tool\b": "Multi-Tool",
        r"\b(bloodline|hurston|micro\s*tech)[' ]+(?=multi[- ]?tool\b)": r"\1 ",
        r"\bmax\s*(?:lit|lift|ift)\b": "MaxLift",
        r"\bkilshot\b": "Killshot",
        r"\bkillshot\b": "Killshot",
        r"\brrie\b": "Rifle",
        r"\brifie\b": "Rifle",
        r"\brile\b": "Rifle",
        r"\brfie\b": "Rifle",
        r"\bparalax\b": "Parallax",
        r"\bsorguine\b": "Sanguine",
        r"\bsarguine\b": "Sanguine",
        r"\bcompensatora\b": "Compensator",
        # The thin "III" in the Deadbolt title commonly collapses into "i" or
        # "im" at screen-share resolution. Keep VI untouched so the two cannon
        # variants remain distinguishable.
        r"\b(?:deed|dead)bolti(?:m)?\s*cannon\b": "Deadbolt III Cannon",
        # Stable one-second screen-share distortions from the permanent Medal
        # corpus. Keep these anchored to full titles so unrelated catalog names
        # cannot be silently coerced.
        r"^csp[- ]*6a\s+bacoacx\s*epooe$": "CSP-68L Backpack Epoque",
        r"^cr\s*het\s+aa\s+support$": "Chiron Helmet AA Support",
        r"^vetse\s*heet\s+excutie$": "Venture Helmet Executive",
        r"^getrcon\s+epoe$": "Geist Armor Core Epoque",
        r"^tana\s*fot\s*heinet\s+murray\s+cup$": "Tailwind Flight Helmet Murray Cup",
        r"^exhet\s*rd\s+alert$": "Arden-SL Helmet Red Alert",
        r"^ae\s*les\s+red\s+aert$": "Arden-SL Legs Red Alert",
        r"^aop4\s*as\s*re\s*alert$": "ADP-mk4 Arms Red Alert",
        r"^eckpeck\s+red\s+alert$": "Arden-CL Backpack Red Alert",
        r"^acopo\s+mono?cte$": "Jacopo Monocle",
        r"^ploe\s+sueter\s+tannenbs$": "Piconalia Sweater Tannenbaum",
        r"^p[u]?ner\s*ma[a]?\s*aa\s*suepot$": "Purifier Mask AA Support",
        r"^teewee\s*boets\s*as[o0]\s*en$": "ThermoWeave Boots ASD Edition",
        r"^c[h]?roe\s+one\s+hend\s+gear\s+strker$": "Chrome Dome Head Gear Striker",
        r"^cser[e]?\s+acket$": "Calister Jacket",
        r"^stheorit$": "StarHeart",
        r"^oayssey\s+racg\s+heinet\s+aipha$": "Odyssey II Racing Helmet Alpha",
        r"^neloi\s+boot\s*and\s*pants\s+striker$": "Navoi Boot and Pants Striker",
        r"^bocat\s+bonber\s+cket\s+netsen$": "Bobcat Bomber Jacket Nelson",
        r"^bcat\s+boner\s+icket\s+tarne$": "Bobcat Bomber Jacket Tarmac",
        r"^deadbolt\s*m\s+cannon$": "Deadbolt III Cannon",
        r"^torrenti\s*module$": "Torrent II Module",
        r"^uo\s*m\s*cargo\s+pushie$": "Tumbril Cargo Plushie",
        r"^uminala\s+s5\s*coin$": "Luminalia '55 Coin",
        r"^abml\s*cargo\s+plushie$": "T.A.B.A. Cargo Plushie",
        r"^redi\s*make\s+ltem\s+fabricator\s+aa\s+support$": "RediMake Item Fabricator AA Support",
        r"\barrowh(?:e|c)ad\b": "Arrowhead",
        r"\bexecutlve\b": "Executive",
        r"\bsnlper\b": "Sniper",
    }
    for pattern, replacement in replacements.items():
        value = re.sub(pattern, replacement, value, flags=re.IGNORECASE)
    value = re.sub(r'(\w)\s+"(?=\s|$)', r'\1"', value)
    return " ".join(value.split())


def _strip_inventory_item_prefix(value: str) -> str:
    """Remove the optional in-game class/grade prefix before catalog matching."""
    return re.sub(
        r"^\s*[a-z]{2,5}\s*(?:[/|\\]\s*)?\d+\s*[/|\\]\s*[a-z]\s*",
        "",
        str(value or ""),
        count=1,
        flags=re.IGNORECASE,
    ).strip()


def _inventory_is_noisy_header(value: str) -> bool:
    normalized = re.sub(r"[^a-z]", "", value.casefold())
    if "access" in normalized or "stor" in normalized and "acce" in normalized:
        return True
    if normalized.endswith("rccese") or normalized.endswith("rccece"):
        return True
    return False


def _inventory_item_from_ocr_line(
    line: str,
    default_location: str,
    default_category: str | None,
) -> dict[str, Any] | None:
    cleaned = _clean_inventory_ocr_line(line)
    if len(cleaned) < 3:
        return None
    blocked = {
        "inventory",
        "local inventory",
        "external inventory",
        "personal inventory",
        "filter",
        "search",
        "category",
        "location",
        "station",
        "quantity",
        "qty",
        "name",
        "items",
    }
    if cleaned.casefold() in blocked:
        return None

    quantity = 1.0
    quantity_match = re.search(r"(?:^|\s)(?:x\s*)?(\d+(?:\.\d+)?)(?:\s*x)?$", cleaned, flags=re.IGNORECASE)
    if quantity_match and quantity_match.start(1) > 0:
        quantity = float(quantity_match.group(1))
        cleaned = cleaned[: quantity_match.start(0)].strip(" -.")
    prefix_match = re.match(r"^(?:x\s*)?(\d+(?:\.\d+)?)\s+(.+)$", cleaned, flags=re.IGNORECASE)
    if prefix_match:
        quantity = float(prefix_match.group(1))
        cleaned = prefix_match.group(2).strip(" -.")

    if len(cleaned) < 3 or cleaned.casefold() in blocked:
        return None
    return {
        "name": cleaned,
        "category": default_category,
        "item_type": None,
        "item_size": None,
        "location": default_location,
        "quantity": quantity,
        "quality": None,
        "volume_scu": None,
        "notes": "Imported from screen capture",
    }


def _clean_blueprint_ocr_line(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9'’+./() -]", " ", value)
    value = " ".join(value.split()).strip(" -.")
    blocked = {"owned", "blueprint", "blueprints", "craft", "category", "search", "filter"}
    if _normalize_text(value) in blocked:
        return ""
    return value


def _blueprint_match_confidence(candidate: str, blueprint_name: str) -> float:
    candidate_norm = _normalize_text(candidate)
    blueprint_norm = _normalize_text(blueprint_name)
    if not candidate_norm or not blueprint_norm:
        return 0
    if candidate_norm == blueprint_norm:
        return 1
    if blueprint_norm in candidate_norm:
        return min(0.95, len(blueprint_norm) / max(len(candidate_norm), 1) + 0.25)
    if candidate_norm in blueprint_norm:
        return min(0.88, len(candidate_norm) / max(len(blueprint_norm), 1) + 0.2)
    candidate_words = set(candidate_norm.split())
    blueprint_words = set(blueprint_norm.split())
    overlap = len(candidate_words & blueprint_words)
    return overlap / max(len(blueprint_words), 1)


@app.get("/api/items")
async def items(
    query: str | None = None,
    category: str | None = None,
    section: str | None = None,
    size: str | None = None,
    limit: int = Query(default=25, ge=1, le=100),
    page: int = Query(default=1, ge=1),
) -> list[dict[str, Any]]:
    return encode(await state().sources.lookup_items(query, category, section, size, limit, page))


@app.get("/api/items/{item_id}")
async def item(item_id: int) -> dict[str, Any]:
    result = await state().sources.lookup_item_by_id(item_id)
    if result is None:
        not_found(f"No item found for id {item_id}.")
    return encode(result)


@app.get("/api/autocomplete/items")
async def autocomplete_items(query: str = "") -> list[str]:
    return await state().sources.autocomplete_items(query)


@app.get("/api/trade/routes")
async def trade_routes(
    starting_point: str,
    ship: str = "Ironclad Assault",
    investment: float = Query(default=1_000_000, gt=0),
    cargo_capacity_scu: float | None = Query(default=None, gt=0),
    max_stops: int = Query(default=5, ge=2, le=5),
    stay_system: str | None = None,
    circular_only: bool = True,
) -> dict[str, Any]:
    capacity = cargo_capacity_scu
    if capacity is None:
        ship_result = await state().sources.lookup_ship(ship)
        capacity = ship_result.cargo_capacity if ship_result and ship_result.cargo_capacity else None
    if capacity is None:
        raise HTTPException(status_code=422, detail="Cargo capacity is required when the ship is unknown.")
    result = await state().sources.lookup_trade_routes(
        ship,
        capacity,
        starting_point,
        investment,
        max_stops,
        stay_system,
        circular_only,
    )
    if result is None or not result.legs:
        not_found("No profitable circular route found.")
    return encode(result)


@app.get("/api/missions")
async def missions(
    query: str | None = None,
    region: str | None = None,
    contractor: str | None = None,
    reputation_level: str | None = None,
    mission_type: str | None = None,
    limit: int = Query(default=50, ge=1, le=100),
    page: int = Query(default=1, ge=1),
) -> list[dict[str, Any]]:
    return encode(await state().sources.lookup_missions(
        query=query, region=region, contractor=contractor,
        reputation_level=reputation_level, mission_type=mission_type,
        limit=limit, page=page,
    ))


@app.get("/api/missions/facets")
async def mission_facets() -> dict[str, list[str]]:
    contractors, reputation_levels, mission_types, systems = await asyncio.gather(
        state().sources.autocomplete_missions("contractor", "", 10000),
        state().sources.autocomplete_missions("reputation_level", "", 10000),
        state().sources.autocomplete_missions("mission_type", "", 10000),
        state().sources.autocomplete_missions("region", "", 10000),
    )
    return {
        "contractors": contractors,
        "reputation_levels": reputation_levels,
        "mission_types": mission_types,
        "systems": systems,
    }


@app.get("/api/autocomplete/missions")
async def autocomplete_missions(filter_name: str = "name", query: str = "") -> list[str]:
    return await state().sources.autocomplete_missions(filter_name, query)


@app.get("/api/autocomplete/trade-locations")
async def autocomplete_trade_locations(query: str = "") -> list[str]:
    return await state().sources.autocomplete_trade_locations(query)


@app.get("/api/exec/status")
async def exec_status() -> dict[str, Any]:
    source_cycle_start = await fetch_exec_cycle_start_unix(state().settings.http_timeout_seconds)
    source_status = calculate_exec_hangar_status(source_cycle_start)
    override = await state().cache.get(EXEC_OVERRIDE_CACHE_KEY)
    active_status = source_status
    if isinstance(override, dict) and isinstance(override.get("cycle_start_unix"), int):
        active_status = calculate_exec_hangar_status(override["cycle_start_unix"])
    return {
        "source": encode(source_status),
        "active": encode(active_status),
        "override": override,
    }


@app.post("/api/exec/override", dependencies=[Depends(require_change_admin)])
async def set_exec_override(payload_request: ExecOverrideRequest, request: Request) -> dict[str, Any]:
    cycle_start = calculate_cycle_start_from_phase(payload_request.phase, payload_request.remaining_minutes)
    payload = {
        "cycle_start_unix": cycle_start,
        "phase": payload_request.phase,
        "remaining_minutes": payload_request.remaining_minutes,
        "corrected_by": payload_request.corrected_by,
        "User": _website_audit_user(current_user_from_request(request, state().settings)),
        "created_at": int(time.time()),
    }
    await state().cache.set(EXEC_OVERRIDE_CACHE_KEY, payload, 315360000)
    await state().cache.add_audit_event("Website Executive Hangar Override Set", payload, "timers")
    return {"status": "saved", "override": payload}


@app.delete("/api/exec/override", dependencies=[Depends(require_change_admin)])
async def clear_exec_override(request: Request) -> dict[str, str]:
    await state().cache.delete(EXEC_OVERRIDE_CACHE_KEY)
    await state().cache.add_audit_event(
        "Website Executive Hangar Override Cleared",
        {"Source": "Website", "User": _website_audit_user(current_user_from_request(request, state().settings))},
        "timers",
    )
    return {"status": "cleared"}


@app.get("/api/cz/timers")
async def cz_timers() -> dict[str, Any]:
    return {
        "definitions": CZ_TIMER_DEFINITIONS,
        "timers": await get_cz_dashboard_timers(state().cache),
    }


@app.post("/api/cz/timers", dependencies=[Depends(require_change_admin)])
async def start_cz_timer(payload: CZTimerRequest, request: Request) -> dict[str, Any]:
    if payload.timer not in CZ_TIMER_DEFINITIONS:
        raise HTTPException(status_code=422, detail="Unknown timer.")
    timers = await get_cz_dashboard_timers(state().cache)
    label, duration = CZ_TIMER_DEFINITIONS[payload.timer]
    timers[payload.timer] = {
        "label": label,
        "ends_at": calculate_countdown_end_unix(duration, payload.started_minutes_ago),
        "duration_seconds": duration,
    }
    await state().cache.set(CZ_TIMERS_CACHE_KEY, timers, 315360000)
    await state().cache.add_audit_event(
        "Website CZ Timer Started",
        {"Timer": label, "User": _website_audit_user(current_user_from_request(request, state().settings))},
        "timers",
    )
    return {"status": "saved", "timers": timers}


@app.delete("/api/cz/timers/{timer}", dependencies=[Depends(require_change_admin)])
async def reset_cz_timer(timer: str, request: Request) -> dict[str, Any]:
    timers = await get_cz_dashboard_timers(state().cache)
    if timer == "all":
        timers = {}
    else:
        timers.pop(timer, None)
    await state().cache.set(CZ_TIMERS_CACHE_KEY, timers, 315360000)
    await state().cache.add_audit_event(
        "Website CZ Timer Reset",
        {"Timer": timer, "User": _website_audit_user(current_user_from_request(request, state().settings))},
        "timers",
    )
    return {"status": "saved", "timers": timers}


@app.get("/api/audit/recent")
async def audit_recent(
    request: Request,
    limit: int = Query(default=25, ge=1, le=100),
    action_type: str | None = Query(default=None),
    sort: str = Query(default="newest", pattern="^(newest|oldest|action)$"),
    _: None = Depends(require_bot_admin),
) -> list[dict[str, Any]]:
    del request
    if action_type and action_type not in AUDIT_ACTION_TYPES:
        raise HTTPException(status_code=422, detail="Unknown audit action type.")
    return await state().cache.recent_audit_events(limit, action_type, sort)


@app.get("/api/audit/visitors")
async def audit_visitors(_: None = Depends(require_bot_admin)) -> dict[str, Any]:
    return await state().cache.website_visitor_analytics()


@app.get("/api/game-data/status")
async def game_data_status() -> dict[str, Any]:
    snapshot_path = Path(__file__).resolve().parents[1] / "data" / "blueprints_snapshot.json"
    try:
        payload = json.loads(snapshot_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise HTTPException(status_code=503, detail="The deployed game database is unavailable.") from error
    source = payload.get("source") or {}
    return {
        "state": "ready",
        "message": "The hosted website and Discord bot are using this deployed game database.",
        "version": source.get("version") or "Unknown",
        "blueprints": len(payload.get("items") or []),
        "missions": len(payload.get("missions") or []),
        "updated_at": int(snapshot_path.stat().st_mtime),
    }


app.mount("/assets", StaticFiles(directory=WEB_DIR), name="assets")


@app.get("/")
async def index() -> FileResponse:
    return FileResponse(
        WEB_DIR / "index.html",
        headers={"Cache-Control": "no-store, max-age=0"},
    )


@app.get("/rsi-hangar-importer/privacy")
async def rsi_hangar_importer_privacy() -> FileResponse:
    return FileResponse(
        WEB_DIR / "rsi-hangar-importer-privacy.html",
        headers={"Cache-Control": "no-store, max-age=0"},
    )
