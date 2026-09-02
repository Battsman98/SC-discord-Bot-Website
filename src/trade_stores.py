from __future__ import annotations

import csv
import hashlib
import io
import re
from dataclasses import dataclass
from urllib.parse import parse_qs, quote, urlparse


MAX_STORE_SHEET_BYTES = 2 * 1024 * 1024
MAX_STORE_ITEMS = 500


@dataclass(frozen=True)
class StoreInventoryItem:
    name: str
    price: str | None
    quantity: str | None
    quality: str | None
    notes: str | None
    location: str | None = None
    category: str | None = None


@dataclass(frozen=True)
class ParsedStoreInventory:
    items: list[StoreInventoryItem]
    content_hash: str


def google_sheet_csv_url(value: str) -> str:
    """Convert a viewable Google Sheets URL into its CSV export URL."""
    raw = str(value or "").strip()
    parsed = urlparse(raw)
    if parsed.scheme != "https" or parsed.hostname not in {"docs.google.com", "drive.google.com"}:
        raise ValueError("Use an https://docs.google.com Google Sheets sharing link.")
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", parsed.path)
    if not match:
        raise ValueError("That link is not a standard Google Sheets sharing link.")
    sheet_id = match.group(1)
    query = parse_qs(parsed.query)
    fragment = parse_qs(parsed.fragment)
    gid = (query.get("gid") or fragment.get("gid") or ["0"])[0]
    if not str(gid).isdigit():
        gid = "0"
    return f"https://docs.google.com/spreadsheets/d/{quote(sheet_id)}/export?format=csv&gid={gid}"


def parse_store_inventory_csv(data: bytes) -> ParsedStoreInventory:
    if not data:
        raise ValueError("The Google Sheet is empty.")
    if len(data) > MAX_STORE_SHEET_BYTES:
        raise ValueError("The Google Sheet is larger than the 2 MB store limit.")
    text = data.decode("utf-8-sig", errors="replace")
    if "<html" in text[:500].casefold():
        raise ValueError("Google returned a sign-in page. Share the sheet as view-only to anyone with the link.")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("The sheet needs a header row.")
    headers = {_normalize_header(name): name for name in reader.fieldnames if name}
    name_header = _find_header(headers, "item name", "item", "name", "product")
    if name_header is None:
        raise ValueError("The sheet needs an Item Name column.")
    price_header = _find_header(headers, "price auec", "price", "auec", "unit price")
    quantity_header = _find_header(headers, "quantity", "qty", "stock")
    quality_header = _find_header(headers, "quality", "grade")
    notes_header = _find_header(headers, "notes", "details", "description")
    items: list[StoreInventoryItem] = []
    for row in reader:
        name = _clean_cell(row.get(name_header))
        if not name:
            continue
        items.append(
            StoreInventoryItem(
                name=name[:150],
                price=_clean_cell(row.get(price_header)) if price_header else None,
                quantity=_clean_cell(row.get(quantity_header)) if quantity_header else None,
                quality=_clean_cell(row.get(quality_header)) if quality_header else None,
                notes=_clean_cell(row.get(notes_header)) if notes_header else None,
                location=_clean_cell(row.get(_find_header(headers, "location", "station"))) if _find_header(headers, "location", "station") else None,
                category=_clean_cell(row.get(_find_header(headers, "category"))) if _find_header(headers, "category") else None,
            )
        )
        if len(items) > MAX_STORE_ITEMS:
            raise ValueError(f"The sheet contains more than {MAX_STORE_ITEMS} inventory rows.")
    if not items:
        raise ValueError("No inventory rows with an item name were found.")
    digest = hashlib.sha256(data).hexdigest()
    return ParsedStoreInventory(items=items, content_hash=digest)


def parse_store_inventory_xlsx(data: bytes) -> ParsedStoreInventory:
    if not data:
        raise ValueError("The Excel inventory file is empty.")
    if len(data) > 5 * 1024 * 1024:
        raise ValueError("The Excel inventory file is larger than the 5 MB store limit.")
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("The attachment is not a readable .xlsx inventory workbook.") from exc
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    first_row = next(rows, None)
    if not first_row:
        raise ValueError("The Excel inventory file needs a header row.")
    headers = {_normalize_header(value): index for index, value in enumerate(first_row) if value is not None}
    name_index = _find_header_index(headers, "item name", "name", "item", "product")
    if name_index is None:
        raise ValueError("The Excel inventory file needs a Name or Item Name column.")
    quantity_index = _find_header_index(headers, "quantity", "qty", "stock")
    quality_index = _find_header_index(headers, "quality", "grade")
    notes_index = _find_header_index(headers, "notes", "details", "description")
    location_index = _find_header_index(headers, "location", "station")
    category_index = _find_header_index(headers, "category")
    price_indices = [
        _find_header_index(headers, "store price auec", "price auec", "unit price", "price"),
        _find_header_index(headers, "average uex player seller price auec"),
        _find_header_index(headers, "average uex terminal sell price auec"),
    ]
    items: list[StoreInventoryItem] = []
    for row in rows:
        name = _xlsx_cell(row, name_index)
        if not name:
            continue
        price = next((_xlsx_cell(row, index) for index in price_indices if _xlsx_cell(row, index)), None)
        items.append(
            StoreInventoryItem(
                name=name[:150],
                price=price,
                quantity=_xlsx_cell(row, quantity_index),
                quality=_xlsx_cell(row, quality_index),
                notes=_xlsx_cell(row, notes_index),
                location=_xlsx_cell(row, location_index),
                category=_xlsx_cell(row, category_index),
            )
        )
        if len(items) > MAX_STORE_ITEMS:
            raise ValueError(f"The Excel inventory contains more than {MAX_STORE_ITEMS} rows.")
    workbook.close()
    if not items:
        raise ValueError("No inventory rows with an item name were found in the Excel file.")
    return ParsedStoreInventory(items=items, content_hash=hashlib.sha256(data).hexdigest())


def _normalize_header(value: object) -> str:
    return " ".join(re.sub(r"[^a-z0-9]+", " ", str(value or "").casefold()).split())


def _find_header(headers: dict[str, str], *candidates: str) -> str | None:
    for candidate in candidates:
        normalized = _normalize_header(candidate)
        if normalized in headers:
            return headers[normalized]
    return None


def _find_header_index(headers: dict[str, int], *candidates: str) -> int | None:
    for candidate in candidates:
        normalized = _normalize_header(candidate)
        if normalized in headers:
            return headers[normalized]
    return None


def _clean_cell(value: object) -> str | None:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    cleaned = " ".join(str(value or "").strip().split())
    return cleaned[:500] if cleaned else None


def _xlsx_cell(row: tuple, index: int | None) -> str | None:
    if index is None or index >= len(row):
        return None
    return _clean_cell(row[index])
