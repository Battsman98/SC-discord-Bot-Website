import asyncio
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

import src.web as web


def test_inventory_export_filters_categories_blanks_notes_and_adds_sell_prices(monkeypatch) -> None:
    items = [
        {
            "name": "FS-9 LMG",
            "category": "Weapons",
            "item_type": "Primary",
            "item_size": "",
            "location": "Everus Harbor",
            "quantity": 2,
            "quality": None,
            "volume_scu": None,
            "notes": "Do not export this note",
            "updated_at": "2026-08-02T12:00:00Z",
        },
        {
            "name": "Gold",
            "category": "Commodities",
            "item_type": "Ore",
            "item_size": "",
            "location": "Everus Harbor",
            "quantity": 4,
            "quality": None,
            "volume_scu": 1,
            "notes": "Hidden",
            "updated_at": "2026-08-02T12:00:00Z",
        },
    ]

    class Cache:
        async def user_inventory_items(self, *_args):
            return items

    class Sources:
        async def inventory_sell_price_comparison(self, names):
            assert names == ["FS-9 LMG"]
            return {"fs 9 lmg": {"terminal_average": 125.5, "player_average": 500}}

    monkeypatch.setattr(web, "state", lambda: SimpleNamespace(cache=Cache(), sources=Sources()))
    response = asyncio.run(
        web.export_my_inventory(
            location=None,
            category=["Weapons"],
            query=None,
            sort_by="location",
            selling=True,
            user=SimpleNamespace(id=7),
        )
    )

    sheet = load_workbook(BytesIO(response.body), data_only=False).active
    headers = [cell.value for cell in sheet[1]]
    assert "Updated At" not in headers
    assert headers[:6] == ["Location", "Category", "Item Type", "Quantity", "Name", "Size"]
    assert "Selected Sell Price (aUEC)" not in headers
    assert headers[-4:] == [
        "Average UEX Terminal Sell Price (aUEC)",
        "Average UEX Player Seller Price (aUEC)",
        "Price Source",
        "Estimated Sell Total (aUEC)",
    ]
    assert sheet.max_row == 2
    assert sheet["D2"].value == 2
    assert sheet["E2"].value == "FS-9 LMG"
    assert sheet["F2"].value is None
    assert sheet["I2"].value is None
    assert sheet["J2"].value == 125.5
    assert sheet["K2"].value == 500
    assert sheet["L2"].value == "UEX Player Marketplace"
    assert sheet["M2"].value == 1000
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:M2"
